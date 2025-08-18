# v021 — Журнал задач с удалением, фильтром, статистикой и стильным Excel

import os
import sys
import time
import datetime
import signal
import re
from urllib.parse import urlparse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter
from colorama import init, Fore, Style

init(autoreset=True)

# ======================
# Настройки
# ======================

def get_log_dir():
    desktop_dir = os.path.join(os.path.expanduser("~"), "Desktop")
    today = datetime.date.today()
    return os.path.join(desktop_dir, "logs", str(today.year), f"{today.month:02d}", f"{today.day:02d}")

LOG_DIR = get_log_dir()
os.makedirs(LOG_DIR, exist_ok=True)

tasks = []  # элементы: {"task","link","time_str","minutes","hours_hundredths"}
start_time = None

# ======================
# Вспомогательные
# ======================

def signal_handler(sig, frame):
    print(Fore.YELLOW + "\nОбнаружено Ctrl+C! Сохраняем данные...")
    save_backup()
    save_excel()
    print(Fore.MAGENTA + "Данные сохранены. Программа завершена.")
    sys.exit(0)

signal.signal(signal.SIGINT, signal_handler)

def parse_domain(url):
    """Вернуть домен из URL (для отображения)"""
    try:
        parsed = urlparse(url)
        return parsed.netloc or url
    except:
        return url

def link_clickable(url, text=None):
    """OSC 8 кликабельная ссылка в терминале (работает в современных консолях)"""
    if not url:
        return ""
    if text is None:
        text = url
    return f"\033]8;;{url}\033\\{text}\033]8;;\033\\"

def key_for_group(t):
    """Ключ группировки: ссылка если есть, иначе название (lower)."""
    if t.get("link"):
        return t["link"].strip().lower()
    return t["task"].strip().lower()

def load_backup():
    today = datetime.date.today().strftime("%Y-%m-%d")
    log_file = os.path.join(LOG_DIR, f"{today}.txt")
    if os.path.exists(log_file):
        with open(log_file, "r", encoding="utf-8") as f:
            for line in f:
                try:
                    task, link, time_str, hours_hundredths = line.strip().split(" | ")
                    if time_str == "<1 минуты":
                        minutes = 0.5
                    else:
                        mins, secs = map(int, time_str.split(":"))
                        minutes = mins + secs / 60
                    tasks.append({
                        "task": task,
                        "link": link,
                        "time_str": time_str,
                        "minutes": round(minutes, 2),
                        "hours_hundredths": float(hours_hundredths)
                    })
                except Exception as e:
                    print(Fore.RED + f"Ошибка чтения backup: {e}")
        if tasks:
            print(Fore.GREEN + f"Загружено {len(tasks)} задач из backup: {log_file}")

def clear_console():
    os.system("cls" if os.name == "nt" else "clear")

def save_backup():
    today = datetime.date.today().strftime("%Y-%m-%d")
    log_file = os.path.join(LOG_DIR, f"{today}.txt")
    with open(log_file, "w", encoding="utf-8") as f:
        for t in tasks:
            f.write(f"{t['task']} | {t['link']} | {t['time_str']} | {t['hours_hundredths']}\n")

def group_tasks():
    """
    Объединяет задачи:
      - по ссылке (если есть)
      - по названию (в любом случае)
    Возвращает список агрегатов, где count > 1
    """
    grouped = {}

    for t in tasks:
        # Сначала группировка по ссылке (если есть)
        link_key = t["link"].strip().lower() if t.get("link") else None
        if link_key:
            if link_key not in grouped:
                grouped[link_key] = {
                    "task": t["task"],
                    "link": t["link"],
                    "minutes": 0.0,
                    "hours_hundredths": 0.0,
                    "count": 0,
                    "key": link_key
                }
            grouped[link_key]["minutes"] += t["minutes"]
            grouped[link_key]["hours_hundredths"] += t["hours_hundredths"]
            grouped[link_key]["count"] += 1

        # Группировка по названию
        name_key = t["task"].strip().lower()
        if name_key not in grouped:
            grouped[name_key] = {
                "task": t["task"],
                "link": t["link"],
                "minutes": 0.0,
                "hours_hundredths": 0.0,
                "count": 0,
                "key": name_key
            }
        grouped[name_key]["minutes"] += t["minutes"]
        grouped[name_key]["hours_hundredths"] += t["hours_hundredths"]
        grouped[name_key]["count"] += 1

    # Оставляем только группы с count > 1
    return [g for g in grouped.values() if g["count"] > 1]


# ======================
# Excel стилизация
# ======================

def style_tasks_and_summary(ws, n_tasks_rows, grouped, group_key_to_rows):
    """
    Стилизация листа:
    - автоширина
    - границы, выравнивание, жирная шапка
    - зебра для обычных строк
    - автофильтр, фиксация шапки
    - выделение строки ИТОГО
    - выделение блоков сводки
    - подсветка повторов в основной таблице и в сводке одинаковыми цветами
    """
    thin = Border(left=Side(style='thin'),
                  right=Side(style='thin'),
                  top=Side(style='thin'),
                  bottom=Side(style='thin'))

    # Палитра мягких цветов для групп (без конфликтов с шапкой)
    palette = [
        "FFF2CC", "E2EFDA", "DDEBF7", "FCE4D6", "EDEDED",
        "CCE5FF", "FFD966", "C6E0B4", "D9E1F2", "FFEB9C",
        "E7E6E6", "D5E8D4", "D0CECE", "F8CBAD", "C9DAF8"
    ]

    def color_for_key(k):
        if not k:
            return None
        idx = abs(hash(k)) % len(palette)
        return PatternFill(start_color=palette[idx], end_color=palette[idx], fill_type="solid")

    max_row = ws.max_row
    max_col = ws.max_column

    # Шапка
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for col in range(1, max_col + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(vertical="center", horizontal="left")
        c.fill = header_fill
        c.border = thin

    # Данные задач: зебра + границы + возможная подсветка группы
    for r in range(2, 2 + n_tasks_rows - 1 if n_tasks_rows > 0 else 1):
        # r — текущая строка с данными
        # зебра по умолчанию
        zebra = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid") if (r % 2 == 0) else None

        # применяем фон по группам (перекрывает зебру)
        # выясняем ключ группы по строке
        if r in group_key_to_rows["row_to_key"]:
            k = group_key_to_rows["row_to_key"][r]
            fill = color_for_key(k)
        else:
            fill = zebra

        for col in range(1, max_col + 1):
            cell = ws.cell(row=r, column=col)
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin

    # ИТОГО — строка после пустой строки
    # ищем строку со значением "ИТОГО" в колонке A
    total_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "ИТОГО":
            total_row = r
            break
    if total_row:
        for col in range(1, max_col + 1):
            c = ws.cell(row=total_row, column=col)
            c.font = Font(bold=True)
            c.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            c.border = thin

    # СВОДКА — заголовок и строки ниже
    summary_header_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "СВОДКА (объединено по совпадению ссылки или названия)":
            summary_header_row = r
            break

    if summary_header_row:
        # Заголовок сводки
        for col in range(1, max_col + 1):
            c = ws.cell(row=summary_header_row, column=col)
            c.font = Font(bold=True)
            c.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            c.border = thin

        # Строки сводки: подсветить каждой группе свой цвет
        r = summary_header_row + 1
        while r <= ws.max_row:
            row_task = ws.cell(row=r, column=1).value
            row_link_domain = ws.cell(row=r, column=2).value
            if row_task is None and row_link_domain is None:
                r += 1
                continue
            # попытка определить ключ по сводке:
            # если ссылка есть в исходных данных, лучше картировать по ссылке, иначе по названию
            # но в сводке второй столбец — домен (не сам URL). Ключ у нас — полная ссылка.
            # поэтому для цвета используем сохранённый словарь group_key_to_rows["key_list"] из аргумента 'grouped'.
            # Создадим карту: task->key и domain->key (по grouped)
            r += 1

        # Проставим цвета по key из grouped:
        key_to_fill = {g["key"]: color_for_key(g["key"]) for g in grouped}
        # Найдём диапазон сводки (после строки summary_header_row до конца/пустой строки)
        sr = summary_header_row + 1
        while sr <= ws.max_row:
            a = ws.cell(row=sr, column=1).value
            b = ws.cell(row=sr, column=2).value
            c3 = ws.cell(row=sr, column=3).value
            if a is None and b is None and c3 is None:
                break
            # Попробуем найти соответствующий key:
            # Пройдём по grouped и найдём совпадение по task или по домену
            matched_key = None
            for g in grouped:
                if a and a == g["task"]:
                    matched_key = g["key"]
                    break
                if b and b == parse_domain(g["link"] or ""):
                    matched_key = g["key"]
                    break
            if matched_key and key_to_fill.get(matched_key):
                fill = key_to_fill[matched_key]
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=sr, column=col)
                    cell.fill = fill
                    cell.border = thin
            sr += 1

    # Автофильтр и фиксация шапки для основной таблицы
    if n_tasks_rows > 0:
        ws.auto_filter.ref = f"A1:D{1 + n_tasks_rows}"
    ws.freeze_panes = "A2"

    # Колонтитулы/выравнивание + границы для всех ячеек
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=max_col):
        for cell in row:
            if not cell.border:
                cell.border = thin
            if not cell.alignment:
                cell.alignment = Alignment(vertical="center")

    # Автоширина столбцов
    for col in range(1, max_col + 1):
        column = get_column_letter(col)
        max_len = 0
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[column].width = max(10, min(60, max_len + 2))

# v022 — Журнал задач с удалением, фильтром, статистикой и Excel со сводкой сразу после ИТОГО

# ... (весь предыдущий код без изменений до функции save_excel) ...

def save_excel():
    today = datetime.date.today().strftime("%Y-%m-%d")
    file_name = os.path.join(LOG_DIR, f"{today}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tasks"

    # Шапка
    ws.append(["Задача", "Ссылка (домен)", "Время (мин:сек)", "Время (часы в сотых)"])

    total_minutes = 0.0
    total_hours_hundredths = 0.0

    # Подготовка карт для подсветки дублей в основной таблице
    keys = [key_for_group(t) for t in tasks]
    dup_keys = {k for k, c in Counter(keys).items() if c > 1}
    row_to_key = {}

    # Данные
    for t in tasks:
        domain = parse_domain(t['link']) if t['link'] else ""
        row_index = ws.max_row + 1
        ws.cell(row=row_index, column=1, value=t['task'])
        if t['link']:
            link_cell = ws.cell(row=row_index, column=2, value=domain)
            link_cell.hyperlink = t['link']
            link_cell.style = "Hyperlink"
        else:
            ws.cell(row=row_index, column=2, value="")
        ws.cell(row=row_index, column=3, value=t['time_str'])
        ws.cell(row=row_index, column=4, value=t['hours_hundredths'])

        k = key_for_group(t)
        if k in dup_keys:
            row_to_key[row_index] = k

        total_minutes += t['minutes']
        total_hours_hundredths += t['hours_hundredths']

    n_tasks_rows = len(tasks) + 1  # включая шапку

    # Пустая строка и ИТОГО
    ws.append([])
    ws.append(["ИТОГО", "", f"{round(total_minutes, 2)} мин", f"{round(total_hours_hundredths, 2)} ч"])

    # СВОДКА сразу после ИТОГО
    grouped = group_tasks()
    if grouped:
        ws.append([])
        ws.append(["СВОДКА (объединено по совпадению ссылки или названия)"])
        for gt in grouped:
            ws.append([
                gt["task"],
                parse_domain(gt["link"]) if gt["link"] else "",
                f"{round(gt['minutes'], 2)} мин",
                f"{round(gt['hours_hundredths'], 2)} ч"
            ])

    # Стилизация листа
    style_tasks_and_summary(
        ws=ws,
        n_tasks_rows=n_tasks_rows,
        grouped=grouped,
        group_key_to_rows={"row_to_key": row_to_key}
    )

    try:
        wb.save(file_name)
        print(Fore.GREEN + f"Excel-отчёт сохранён: {file_name}")
    except Exception as e:
        print(Fore.RED + f"Ошибка при сохранении Excel: {e}")


def display_tasks(filter_text=None):
    clear_console()
    print(Fore.CYAN + "=== Журнал задач ===")
    for idx, t in enumerate(tasks, start=1):
        # фильтр по подстроке в названии или ссылке
        if filter_text:
            ft = filter_text.lower()
            if ft not in t['task'].lower() and ft not in t['link'].lower():
                continue
        domain = parse_domain(t['link']) if t['link'] else ""
        link_part = f" [{link_clickable(t['link'], domain)}]" if t['link'] else ""
        print(
            Fore.YELLOW + f"{idx}. {t['task']}{link_part}" +
            Fore.WHITE + f" — {t['time_str']} ({t['hours_hundredths']} ч)"
        )

def edit_task(index):
    try:
        t = tasks[index]
        print(Fore.CYAN + f"Редактирование задачи #{index+1}")
        new_task = input(f"Текст задачи ({t['task']}): ").strip() or t['task']
        new_link = input(f"Ссылка ({t['link']}): ").strip() or t['link']
        new_time = input(f"Время (минуты, текущее {t['minutes']}): ").strip()

        if new_time:
            try:
                minutes = float(new_time)
            except:
                print(Fore.RED + "Неверный ввод времени, оставлено старое значение")
                minutes = t['minutes']
        else:
            minutes = t['minutes']

        tasks[index] = {
            "task": new_task,
            "link": new_link,
            "minutes": round(minutes, 2),
            "hours_hundredths": round(minutes / 60, 2),
            "time_str": "<1 минуты" if minutes < 1 else f"{int(minutes)}:{int((minutes%1)*60):02d}"
        }
        save_backup()
        display_tasks()
    except IndexError:
        print(Fore.RED + "Нет задачи с таким номером.")

def delete_task(index):
    try:
        t = tasks[index]
        print(Fore.CYAN + f"\nВы собираетесь удалить задачу #{index+1}:")
        print(Fore.YELLOW + f"Задача: {t['task']}")
        print(Fore.YELLOW + f"Ссылка: {t['link']}")
        print(Fore.YELLOW + f"Время: {t['time_str']} ({t['hours_hundredths']} ч)")

        confirm = input(Fore.RED + "Вы уверены? (y/n): ").strip().lower()
        if confirm == "y":
            tasks.pop(index)
            save_backup()
            print(Fore.GREEN + "Задача удалена.")
            display_tasks()
        else:
            print(Fore.MAGENTA + "Удаление отменено.")
    except IndexError:
        print(Fore.RED + "Нет задачи с таким номером.")

def delete_all_tasks():
    confirm = input(Fore.RED + "Вы уверены, что хотите удалить ВСЕ задачи? (y/n): ").strip().lower()
    if confirm == "y":
        tasks.clear()
        save_backup()
        print(Fore.GREEN + "Все задачи удалены.")
        display_tasks()
    else:
        print(Fore.MAGENTA + "Удаление отменено.")

def show_stats():
    print(Fore.CYAN + "\n=== Статистика ===")
    total_tasks = len(tasks)
    total_minutes = round(sum(t['minutes'] for t in tasks), 2)
    total_hours = round(sum(t['hours_hundredths'] for t in tasks), 2)

    print(Fore.YELLOW + f"Всего задач: {total_tasks}")
    print(Fore.YELLOW + f"Общее время: {total_minutes} мин / {total_hours} ч")

    # топ-3 ссылок по времени
    link_minutes = {}
    for t in tasks:
        if t['link']:
            d = parse_domain(t['link'])
            link_minutes[d] = link_minutes.get(d, 0) + t['minutes']
    if link_minutes:
        print(Fore.GREEN + "\nТоп-3 ссылок (по времени):")
        for domain, mins in sorted(link_minutes.items(), key=lambda x: x[1], reverse=True)[:3]:
            print(Fore.WHITE + f" - {domain}: {round(mins, 2)} мин")

    # топ-3 задач по времени
    task_minutes = {}
    for t in tasks:
        task_minutes[t['task']] = task_minutes.get(t['task'], 0) + t['minutes']
    if task_minutes:
        print(Fore.GREEN + "\nТоп-3 задач (по времени):")
        for name, mins in sorted(task_minutes.items(), key=lambda x: x[1], reverse=True)[:3]:
            print(Fore.WHITE + f" - {name}: {round(mins, 2)} мин")

# ======================
# Запуск
# ======================

load_backup()
display_tasks()

print(Fore.GREEN +
      "Вводите задачи (! — уже выполненная).\n"
      "Команды:\n"
      " :h — завершить день (сохранить Excel и выйти)\n"
      " :s — сохранить Excel\n"
      " :q — выйти (с сохранением Excel)\n"
      " :e1 — редактировать задачу №1\n"
      " :d1 — удалить задачу №1 (с подтверждением)\n"
      " :da — удалить все задачи (с подтверждением)\n"
      " :stats — показать статистику\n"
)

while True:
    try:
        task_input = input(Fore.CYAN + "\nВведите задачу: ").strip()

        # Редактирование
        if re.match(r"^:e\d+$", task_input) or re.match(r"^:edit\d+$", task_input):
            idx = int(re.findall(r"\d+", task_input)[0]) - 1
            edit_task(idx)
            continue

        # Удаление одной
        if re.match(r"^:d\d+$", task_input) or re.match(r"^:del\d+$", task_input):
            idx = int(re.findall(r"\d+", task_input)[0]) - 1
            delete_task(idx)
            continue

        # Удаление всех
        if task_input.lower() in [":da", ":delall"]:
            delete_all_tasks()
            continue

        # Фильтр
        if task_input.lower().startswith(":f "):
            query = task_input[3:].strip()
            display_tasks(filter_text=query)
            continue

        # Статистика
        if task_input.lower() == ":stats":
            show_stats()
            continue

        # Выход и сохранение
        if task_input.lower() in [":home", ":h", ":q", ":quit", ":exit"]:
            save_excel()
            print(Fore.MAGENTA + "День завершён.")
            break
        elif task_input.lower() in [":s", ":save"]:
            save_excel()
            continue
        elif not task_input:
            print(Fore.RED + "Пожалуйста, введите задачу или команду.")
            continue

        # Добавление задачи
        is_completed = task_input.startswith("!")
        task_text = task_input[1:] if is_completed else task_input
        link = input(Fore.CYAN + "Введите ссылку (если есть): ").strip()

        if is_completed:
            while True:
                try:
                    minutes = float(input(Fore.CYAN + "Время (в минутах): ").strip())
                    break
                except ValueError:
                    print(Fore.RED + "Введите число!")
        else:
            start_time = time.time()
            input(Fore.CYAN + "Нажмите Enter, когда задача завершена...")
            elapsed = time.time() - start_time
            minutes = elapsed / 60

        hours_hundredths = round(minutes / 60, 2)
        time_str = "<1 минуты" if minutes < 1 else f"{int(minutes)}:{int((minutes%1)*60):02d}"

        tasks.append({
            "task": task_text,
            "link": link,
            "time_str": time_str,
            "minutes": round(minutes, 2),
            "hours_hundredths": hours_hundredths
        })

        save_backup()
        display_tasks()

    except KeyboardInterrupt:
        signal_handler(None, None)
