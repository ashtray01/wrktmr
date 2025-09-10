# v022.5 — Журнал задач с GUI, живыми таймерами, Markdown-ссылками, фиксированной высотой задач, выбором папки сохранения и временным хранилищем

import os
import sys
import time
import datetime
import signal
import re
import tempfile
import ctypes
import uuid
import PyQt6
from urllib.parse import urlparse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter
from colorama import init, Fore, Style

# ======================
# Темная тема заголовка окна для Windows 10/11
# ======================

def set_dark_title_bar_qt(window):
    """Установка тёмного заголовка окна в Windows 10/11 для PyQt6"""
    if sys.platform != "win32":
        return

    try:
        # Получаем HWND окна
        hwnd = int(window.winId())

        # Включаем тёмный режим заголовка
        # DWMWA_USE_IMMERSIVE_DARK_MODE = 20
        dark_mode_attribute = 20
        value = ctypes.c_int(1)
        ctypes.windll.dwmapi.DwmSetWindowAttribute(
            hwnd,
            dark_mode_attribute,
            ctypes.byref(value),
            ctypes.sizeof(value)
        )
    except Exception as e:
        print(f"[Тёмный заголовок] Не удалось применить: {e}")

# GUI
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
                             QTextEdit, QLineEdit, QPushButton, QLabel, QScrollArea, 
                             QFrame, QMessageBox, QFileDialog, QComboBox, QDialog, 
                             QDialogButtonBox, QFormLayout, QSpinBox, QDoubleSpinBox)
from PyQt6.QtCore import Qt, QTimer, QUrl, QTime, QSize
from PyQt6.QtGui import QFont, QDesktopServices, QColor, QTextCharFormat, QTextCursor, QPalette

init(autoreset=True)

# ======================
# Настройки
# ======================

def get_temp_log_dir():
    """Временная папка для логов (сохраняется только до перезапуска)"""
    temp_dir = tempfile.gettempdir()
    today = datetime.date.today()
    return os.path.join(temp_dir, "task_journal", str(today.year), f"{today.month:02d}", f"{today.day:02d}")

# Глобальные настройки
LOG_DIR = get_temp_log_dir()
os.makedirs(LOG_DIR, exist_ok=True)

# Путь для сохранения Excel — по умолчанию Desktop, но можно выбрать
SAVE_DIR = os.path.join(os.path.expanduser("~"), "Desktop")

tasks = []  # элементы: {"task","link","time_str","minutes","hours_hundredths", "timer_start" (опционально)}
active_tasks_file = os.path.join(LOG_DIR, "active_tasks.tmp")
active_save_timer = None  # Таймер для периодического сохранения
current_theme = "dark"  # или "light"

# ======================
# Вспомогательные
# ======================

def signal_handler(sig, frame):
    print(Fore.YELLOW + "\nОбнаружено Ctrl+C! Сохраняем данные...")
    save_backup()
    print(Fore.MAGENTA + "Данные сохранены. Программа завершена.")
    sys.exit(0)

signal.signal(signal.SIGINT, signal_handler)

def parse_domain(url):
    try:
        parsed = urlparse(url)
        return parsed.netloc or url
    except:
        return url
    
def format_start_time(timestamp):
    """Форматирует timestamp в локальное системное время (с часовым поясом, если есть)"""
    try:
        if not isinstance(timestamp, (int, float)):
            return "Начата: время неизвестно"

        # Преобразуем timestamp в datetime с локальной таймзоной системы
        dt = datetime.datetime.fromtimestamp(timestamp)

        # Пытаемся получить имя часового пояса
        tz_name = dt.strftime("%Z")  # Например: "MSK", "CET", "PDT", или пусто
        tz_info = f" ({tz_name})" if tz_name.strip() else ""

        # Форматируем: 14:22 15.04.2025 (MSK)
        formatted = dt.strftime("%H:%M %d.%m.%Y") + tz_info

        return f"Начата: {formatted}"
    except Exception as e:
        print(f"Ошибка форматирования времени: {e}")
        return "Начата: время неизвестно"

def key_for_group(t):
    if t.get("link"):
        return t["link"].strip().lower()
    return t["task"].strip().lower()

def load_backup():
    today = datetime.date.today().strftime("%Y-%m-%d")
    log_file = os.path.join(LOG_DIR, f"{today}.txt")
    if os.path.exists(log_file):
        with open(log_file, "r", encoding="utf-8") as f:
            for line in f:
                try:
                    parts = line.strip().split(" | ")
                    if len(parts) < 4:
                        continue
                    task, link, time_str, hours_hundredths = parts
                    if time_str == "<1 минуты":
                        minutes = 0.5
                    else:
                        if "-" in time_str:
                            minutes = parse_time_range(time_str)
                        else:
                            try:
                                mins, secs = map(int, time_str.split(":"))
                                minutes = mins + secs / 60
                            except:
                                minutes = float(time_str) if time_str.replace('.','',1).isdigit() else 0.0

                    tasks.append({
                        "id": str(uuid.uuid4()),
                        "task": task,
                        "link": link,
                        "time_str": time_str,
                        "minutes": round(minutes, 2),
                        "hours_hundredths": float(hours_hundredths)
                    })

                except Exception as e:
                    print(Fore.RED + f"Ошибка чтения backup: {e}")

        if tasks:
            print(Fore.GREEN + f"Загружено {len(tasks)} задач из backup: {log_file}")

    # ✅ Фильтрация дублей — ДЕЛАЕМ ЭТО В КОНЦЕ, ПОСЛЕ ЗАГРУЗКИ АКТИВНЫХ ЗАДАЧ
    active_ids = {t.get("id") for t in tasks if "timer_start" in t}
    if active_ids:
        # Оставляем активные задачи + завершённые, которых нет среди активных
        tasks[:] = [t for t in tasks if "timer_start" in t or t.get("id") not in active_ids]

def load_active_tasks():
    """Загружает активные задачи из временного файла и возобновляет таймеры"""
    if not os.path.exists(active_tasks_file):
        return

    try:
        with open(active_tasks_file, "r", encoding="utf-8") as f:
            for line in f:
                try:
                    parts = line.strip().split(" | ")
                    if len(parts) < 4:
                        continue
                    task_id, task, link, start_timestamp_str = parts
                    start_timestamp = float(start_timestamp_str)

                    tasks.append({
                        "id": task_id,
                        "task": task,
                        "link": link,
                        "timer_start": start_timestamp,
                        "minutes": 0.0,
                        "hours_hundredths": 0.0,
                        "time_str": "0:00"
                    })
                except Exception as e:
                    print(Fore.RED + f"Ошибка загрузки активной задачи: {e}")

        if tasks:
            active_count = len([t for t in tasks if 'timer_start' in t])
            print(Fore.GREEN + f"Загружено {active_count} активных задач из временного файла.")
    except Exception as e:
        print(Fore.RED + f"Ошибка чтения активных задач: {e}")

def save_backup():
    today = datetime.date.today().strftime("%Y-%m-%d")
    log_file = os.path.join(LOG_DIR, f"{today}.txt")
    os.makedirs(os.path.dirname(log_file), exist_ok=True)
    with open(log_file, "w", encoding="utf-8") as f:
        for t in tasks:
            # 🔥 НЕ сохраняем задачи, которые ещё активны (имеют timer_start)
            if "timer_start" in t:
                continue
            clean_t = {k: v for k, v in t.items() if k != 'timer_start'}
            f.write(f"{clean_t['task']} | {clean_t['link']} | {clean_t['time_str']} | {clean_t['hours_hundredths']}\n")
    #дебаг
    #print(f"[DEBUG] Сохранено {len(tasks)} задач, из них активных: {len([t for t in tasks if 'timer_start' in t])}")

def save_active_tasks():
    """Сохраняет только задачи 'в процессе' (с timer_start) во временный файл, с ID"""
    try:
        active_tasks = [t for t in tasks if "timer_start" in t]
        if not active_tasks:
            if os.path.exists(active_tasks_file):
                os.remove(active_tasks_file)
            return

        with open(active_tasks_file, "w", encoding="utf-8") as f:
            for t in active_tasks:
                task_copy = t.copy()
                start_time = task_copy.pop("timer_start")
                task_id = task_copy.get("id", str(uuid.uuid4()))  # на случай, если id нет
                # Формат: id | task | link | start_timestamp
                f.write(f"{task_id} | {task_copy['task']} | {task_copy['link']} | {start_time}\n")
    #дебаг
    #print(f"[DEBUG] Сохранено {len(tasks)} задач, из них активных: {len([t for t in tasks if 'timer_start' in t])}")


    except Exception as e:
        print(Fore.RED + f"Ошибка сохранения активных задач: {e}")

def remove_active_task(index):
    """Удаляет активную задачу из временного файла по индексу (после завершения)"""
    try:
        if index < 0 or index >= len(tasks):
            return

        task_id = tasks[index].get("id")
        if not task_id:
            return

        # Перечитываем все активные задачи
        if not os.path.exists(active_tasks_file):
            return

        remaining_lines = []
        removed = False

        with open(active_tasks_file, "r", encoding="utf-8") as f:
            for line in f:
                parts = line.strip().split(" | ")
                if len(parts) < 4:  # id | task | link | start_timestamp
                    remaining_lines.append(line.strip())
                    continue

                saved_id, _, _, _ = parts
                if saved_id != task_id:
                    remaining_lines.append(line.strip())
                else:
                    removed = True

        # Перезаписываем файл
        with open(active_tasks_file, "w", encoding="utf-8") as f:
            for line in remaining_lines:
                f.write(line + "\n")

        if not remaining_lines:
            os.remove(active_tasks_file)

        if removed:
            print(Fore.GREEN + f"Активная задача {task_id} удалена из временного файла.")

    except Exception as e:
        print(Fore.RED + f"Ошибка удаления активной задачи: {e}")

def group_tasks():
    grouped = {}
    for t in tasks:
        link_key = t["link"].strip().lower() if t.get("link") else None
        if link_key:
            if link_key not in grouped:
                grouped[link_key] = {
                    "task": t["task"],
                    "link": t["link"],
                    "minutes": 0.0,
                    "hours_hundredths": 0.0,
                    "count": 0,
                    "key": link_key
                }
            grouped[link_key]["minutes"] += t["minutes"]
            grouped[link_key]["hours_hundredths"] += t["hours_hundredths"]
            grouped[link_key]["count"] += 1

        name_key = t["task"].strip().lower()
        if name_key not in grouped:
            grouped[name_key] = {
                "task": t["task"],
                "link": t["link"],
                "minutes": 0.0,
                "hours_hundredths": 0.0,
                "count": 0,
                "key": name_key
            }
        grouped[name_key]["minutes"] += t["minutes"]
        grouped[name_key]["hours_hundredths"] += t["hours_hundredths"]
        grouped[name_key]["count"] += 1

    return [g for g in grouped.values() if g["count"] > 1]

# ======================
# Парсинг времени 13:24-16:39
# ======================

def parse_time_range(time_str):
    try:
        start_str, end_str = time_str.split("-")
        h1, m1 = map(int, start_str.split(":"))
        h2, m2 = map(int, end_str.split(":"))
        start_minutes = h1 * 60 + m1
        end_minutes = h2 * 60 + m2
        if end_minutes < start_minutes:
            end_minutes += 24 * 60
        diff = end_minutes - start_minutes
        return diff
    except:
        return 0.0

# ======================
# Обработка Markdown-ссылок [текст](url)
# ======================

def extract_markdown_links(text):
    pattern = r'\[([^\]]+)\]\(([^)]+)\)'
    matches = re.findall(pattern, text)
    if matches:
        display_text = matches[0][0]
        url = matches[0][1]
        clean_text = re.sub(pattern, display_text, text, count=1)
        return clean_text.strip(), url.strip()
    return text.strip(), ""

# ======================
# Excel стилизация и сохранение
# ======================

def style_tasks_and_summary(ws, n_tasks_rows, grouped, group_key_to_rows):
    thin = Border(left=Side(style='thin'),
                  right=Side(style='thin'),
                  top=Side(style='thin'),
                  bottom=Side(style='thin'))

    palette = [
        "FFF2CC", "E2EFDA", "DDEBF7", "FCE4D6", "EDEDED",
        "CCE5FF", "FFD966", "C6E0B4", "D9E1F2", "FFEB9C",
        "E7E6E6", "D5E8D4", "D0CECE", "F8CBAD", "C9DAF8"
    ]

    def color_for_key(k):
        if not k:
            return None
        idx = abs(hash(k)) % len(palette)
        return PatternFill(start_color=palette[idx], end_color=palette[idx], fill_type="solid")

    max_row = ws.max_row
    max_col = ws.max_column

    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for col in range(1, max_col + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(vertical="center", horizontal="left")
        c.fill = header_fill
        c.border = thin

    for r in range(2, 2 + n_tasks_rows - 1 if n_tasks_rows > 0 else 1):
        zebra = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid") if (r % 2 == 0) else None
        if r in group_key_to_rows["row_to_key"]:
            k = group_key_to_rows["row_to_key"][r]
            fill = color_for_key(k)
        else:
            fill = zebra

        for col in range(1, max_col + 1):
            cell = ws.cell(row=r, column=col)
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin

    total_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "ИТОГО":
            total_row = r
            break
    if total_row:
        for col in range(1, max_col + 1):
            c = ws.cell(row=total_row, column=col)
            c.font = Font(bold=True)
            c.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            c.border = thin

    summary_header_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "СВОДКА (объединено по совпадению ссылки или названия)":
            summary_header_row = r
            break

    if summary_header_row:
        for col in range(1, max_col + 1):
            c = ws.cell(row=summary_header_row, column=col)
            c.font = Font(bold=True)
            c.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            c.border = thin

        sr = summary_header_row + 1
        while sr <= ws.max_row:
            a = ws.cell(row=sr, column=1).value
            b = ws.cell(row=sr, column=2).value
            c3 = ws.cell(row=sr, column=3).value
            if a is None and b is None and c3 is None:
                break
            matched_key = None
            for g in grouped:
                if a and a == g["task"]:
                    matched_key = g["key"]
                    break
                if b and b == parse_domain(g["link"] or ""):
                    matched_key = g["key"]
                    break
            if matched_key:
                fill = color_for_key(matched_key)
                if fill:
                    for col in range(1, max_col + 1):
                        cell = ws.cell(row=sr, column=col)
                        cell.fill = fill
                        cell.border = thin
            sr += 1

    if n_tasks_rows > 0:
        ws.auto_filter.ref = f"A1:D{1 + n_tasks_rows}"
    ws.freeze_panes = "A2"

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=max_col):
        for cell in row:
            if not cell.border:
                cell.border = thin
            if not cell.alignment:
                cell.alignment = Alignment(vertical="center")

    for col in range(1, max_col + 1):
        column = get_column_letter(col)
        max_len = 0
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[column].width = max(10, min(60, max_len + 2))

def save_excel():
    global SAVE_DIR
    today = datetime.date.today().strftime("%Y-%m-%d")
    file_name = os.path.join(SAVE_DIR, f"{today}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tasks"

    ws.append(["Задача", "Ссылка (домен)", "Время (мин:сек)", "Время (часы в сотых)"])

    total_minutes = 0.0
    total_hours_hundredths = 0.0

    keys = [key_for_group(t) for t in tasks]
    dup_keys = {k for k, c in Counter(keys).items() if c > 1}
    row_to_key = {}

    for t in tasks:
        domain = parse_domain(t['link']) if t['link'] else ""
        row_index = ws.max_row + 1
        ws.cell(row=row_index, column=1, value=t['task'])
        if t['link']:
            link_cell = ws.cell(row=row_index, column=2, value=domain)
            link_cell.hyperlink = t['link']
            link_cell.style = "Hyperlink"
        else:
            ws.cell(row=row_index, column=2, value="")
        ws.cell(row=row_index, column=3, value=t['time_str'])
        ws.cell(row=row_index, column=4, value=t['hours_hundredths'])

        k = key_for_group(t)
        if k in dup_keys:
            row_to_key[row_index] = k

        total_minutes += t['minutes']
        total_hours_hundredths += t['hours_hundredths']

    n_tasks_rows = len(tasks) + 1

    ws.append([])
    ws.append(["ИТОГО", "", f"{round(total_minutes, 2)} мин", f"{round(total_hours_hundredths, 2)} ч"])

    grouped = group_tasks()
    if grouped:
        ws.append([])
        ws.append(["СВОДКА (объединено по совпадению ссылки или названия)"])
        for gt in grouped:
            ws.append([
                gt["task"],
                parse_domain(gt["link"]) if gt["link"] else "",
                f"{round(gt['minutes'], 2)} мин",
                f"{round(gt['hours_hundredths'], 2)} ч"
            ])

    style_tasks_and_summary(
        ws=ws,
        n_tasks_rows=n_tasks_rows,
        grouped=grouped,
        group_key_to_rows={"row_to_key": row_to_key}
    )

    try:
        wb.save(file_name)
        print(Fore.GREEN + f"Excel-отчёт сохранён: {file_name}")
        return file_name
    except Exception as e:
        print(Fore.RED + f"Ошибка при сохранении Excel: {e}")
        return None

# ======================
# GUI Классы
# ======================

class TaskItemWidget(QFrame):
    def __init__(self, task_data, index, parent=None):
        super().__init__(parent)
        self.task_data = task_data
        self.index = index
        self.parent_window = parent
        self.timer_label = None
        self.timer = None
        self.setup_ui()
        if "timer_start" in self.task_data:
            self.start_live_timer()

    def setup_ui(self):
        if current_theme == "light":
            self.setStyleSheet("""
                        TaskItemWidget {
                            background-color: white;
                            border-radius: 8px;
                            border: 1px solid #e0e0e0;
                            margin: 4px;
                        }
                    """)
        else:
            self.setStyleSheet("""
                        TaskItemWidget {
                            background-color: #2a2a2a;
                            border-radius: 6px;
                            border: 1px solid #444;
                            margin: 4px;
                        }
                    """)
        self.setFrameShadow(QFrame.Shadow.Raised)
        main_layout = QHBoxLayout()
        main_layout.setContentsMargins(10, 8, 10, 8)
        main_layout.setSpacing(10)

        # Левая часть — описание задачи
        left_layout = QVBoxLayout()
        left_layout.setSpacing(3)

        # Задача
        task_label = QLabel()
        task_text = self.task_data['task']
        task_label.setTextFormat(Qt.TextFormat.RichText)
        task_label.setText(self.format_markdown_links(task_text))
        task_label.setWordWrap(True)
        task_label.setTextInteractionFlags(Qt.TextInteractionFlag.TextBrowserInteraction)
        task_label.linkActivated.connect(self.open_link)
        task_label.setStyleSheet("font-size: 13px;")
        left_layout.addWidget(task_label)

        # Ссылка
        if self.task_data['link']:
            domain = parse_domain(self.task_data['link'])
            link_label = QLabel(f"<a href='{self.task_data['link']}' style='color: #64b5f6; font-size: 11px;'>🔗 {domain}</a>")
            link_label.setTextFormat(Qt.TextFormat.RichText)
            link_label.setTextInteractionFlags(Qt.TextInteractionFlag.TextBrowserInteraction)
            link_label.linkActivated.connect(self.open_link)
            left_layout.addWidget(link_label)

        # Время
        self.time_display = QLabel()
        self.update_time_display()
        self.time_display.setStyleSheet("font-size: 11px; color: #aaa;")
        left_layout.addWidget(self.time_display)

        main_layout.addLayout(left_layout)

        # Правая часть — кнопки (ширина 15% от окна, высота 32px)
        button_layout = QVBoxLayout()
        button_layout.setSpacing(5)
        button_layout.addStretch()

        button_width = int(self.parent_window.width() * 0.2) if self.parent_window else 150  # +50% ширины
        button_height = 32

        if "timer_start" in self.task_data:
            self.finish_btn = QPushButton("⏹️ Завершить задачу")
            self.finish_btn.setFixedSize(button_width, button_height)
            self.finish_btn.clicked.connect(self.confirm_finish_task)
            self.finish_btn.setStyleSheet("""
                QPushButton {
                    background-color: #d32f2f;   /* 🔴 Красный цвет */
                    color: white;
                    border: 1px solid #b71c1c;
                    border-radius: 4px;
                    padding: 4px 8px;
                }
                QPushButton:hover {
                    background-color: #f44336;   /* 🔥 Ярче при наведении */
                }
            """)
            button_layout.addWidget(self.finish_btn)
            self.start_time_label = QLabel()
            self.start_time_label.setFixedWidth(button_width)  # 🔥 Фиксируем ширину как у кнопок
            self.start_time_label.setStyleSheet("font-size: 10px; color: #aaa; padding-top: 5px;")
            self.start_time_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            self.start_time_label.setWordWrap(True)  # 🔥 На случай, если текст не поместится
            self.update_start_time_label()
            button_layout.addWidget(self.start_time_label)
        else:
            edit_btn = QPushButton("✏️ Редактировать")
            edit_btn.setFixedSize(button_width, button_height)
            edit_btn.clicked.connect(self.edit_task)

            delete_btn = QPushButton("🗑️ Удалить")
            delete_btn.setFixedSize(button_width, button_height)
            delete_btn.clicked.connect(self.delete_task)

            button_layout.addWidget(edit_btn)
            button_layout.addWidget(delete_btn)

        button_layout.addStretch()
        main_layout.addLayout(button_layout)

        self.setLayout(main_layout)
        self.setFixedHeight(90)  # Фиксированная компактная высота

    def format_markdown_links(self, text):
        pattern = r'\[([^\]]+)\]\(([^)]+)\)'
        def repl(match):
            display = match.group(1)
            url = match.group(2)
            return f'<a href="{url}" style="color: #64b5f6;">{display}</a>'
        return re.sub(pattern, repl, text)

    def open_link(self, url):
        QDesktopServices.openUrl(QUrl(url))

    def update_time_display(self):
        if "timer_start" in self.task_data:
            elapsed = time.time() - self.task_data['timer_start']
            minutes = elapsed / 60
            
            # Анимация / | \ -
            anim_chars = ['|', '/', '-', '\\']
            if not hasattr(self, '_anim_index'):
                self._anim_index = 0
            else:
                self._anim_index = (self._anim_index + 1) % 4

            anim_char = anim_chars[self._anim_index]
            self.time_display.setText(f"<span style='color: #4caf50;'>⏱️ В процессе ({int(minutes)} мин)</span> {anim_char}")
        else:
            self.time_display.setText(f"⏱️ {self.task_data['time_str']} ({self.task_data['hours_hundredths']} ч)")

    def start_live_timer(self):
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.update_time_display)
        self.timer.start(1000)

    def confirm_finish_task(self):
        reply = QMessageBox.question(
            self,
            "Завершение задачи",
            "Вы уверены, что хотите завершить эту задачу?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.Yes:
            self.finish_task()

    def finish_task(self):
        if "timer_start" not in self.task_data:
            return
        remove_active_task(self.index)

        elapsed = time.time() - self.task_data['timer_start']
        minutes = elapsed / 60
        hours_hundredths = round(minutes / 60, 2)
        time_str = "<1 минуты" if minutes < 1 else f"{int(minutes)}:{int((minutes%1)*60):02d}"

        tasks[self.index].update({
            "minutes": round(minutes, 2),
            "hours_hundredths": hours_hundredths,
            "time_str": time_str
        })
        del tasks[self.index]["timer_start"]

        if self.timer:
            self.timer.stop()

        save_backup()
        self.parent_window.load_tasks_to_ui()

    def edit_task(self):
        self.parent_window.edit_task(self.index)

    def delete_task(self):
        self.parent_window.delete_task(self.index)

    def update_start_time_label(self):
        if hasattr(self, 'start_time_label') and "timer_start" in self.task_data:
            start_timestamp = self.task_data["timer_start"]
            if isinstance(start_timestamp, (int, float)):
                time_str = format_start_time(start_timestamp)
                self.start_time_label.setText(time_str)
                self.start_time_label.show()
            else:
                self.start_time_label.hide()
        elif hasattr(self, 'start_time_label'):
            self.start_time_label.hide()

class EditTaskDialog(QDialog):
    def __init__(self, task_data, parent=None):
        super().__init__(parent)
        self.task_data = task_data
        self.setWindowTitle("✏️ Редактировать задачу")
        self.resize(500, 200)
        self.setup_ui()

    def setup_ui(self):
        layout = QFormLayout()

        self.task_input = QLineEdit(self.task_data.get("task", ""))
        self.link_input = QLineEdit(self.task_data.get("link", ""))
        
        # 🔥 ИСПРАВЛЕНО: В поле времени передаём ЧИСЛО МИНУТ (не time_str!)
        minutes_value = self.task_data.get("minutes", 0.0)
        self.time_input = QLineEdit(str(minutes_value))  # ← Важно! Не time_str, а minutes

        layout.addRow("📝 Задача:", self.task_input)
        layout.addRow("🔗 Ссылка:", self.link_input)
        layout.addRow("⏱️ Время (в минутах):", self.time_input)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)

        self.setLayout(layout)

    def get_data(self):
        task_text = self.task_input.text().strip()
        link = self.link_input.text().strip()
        time_input_text = self.time_input.text().strip()

        minutes = 0.0
        try:
            minutes = float(time_input_text)
        except:
            QMessageBox.warning(self, "Ошибка", "Неверный формат времени. Введите число минут.")
            return None

        hours_hundredths = round(minutes / 60, 2)
        time_str = "<1 минуты" if minutes < 1 else f"{int(minutes)}:{int((minutes%1)*60):02d}"

        return {
            "task": task_text,
            "link": link,
            "time_str": time_str,
            "minutes": round(minutes, 2),
            "hours_hundredths": hours_hundredths
        }


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("⏰ wrktmr v022.7")
        self.setGeometry(100, 100, 900, 700)
        self.tasks_widgets = []
        load_active_tasks()
        load_backup()
        self.init_ui()
        self.apply_theme()
        self.load_tasks_to_ui()
        self.start_active_save_timer()
        self.show()
        set_dark_title_bar_qt(self)
    def start_active_save_timer(self):
        """Запускает таймер для периодического сохранения активных задач"""
        global active_save_timer
        if active_save_timer:
            return

        def save_loop():
            save_active_tasks()
            active_save_timer = QTimer.singleShot(10000, save_loop)  # каждые 10 сек

        save_loop()

    def init_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        # Верхняя панель: тема + путь сохранения
        top_layout = QHBoxLayout()

        self.theme_combo = QComboBox()
        self.theme_combo.addItems(["🌙 Тёмная тема", "☀️ Светлая тема"])
        self.theme_combo.currentIndexChanged.connect(self.toggle_theme)
        top_layout.addWidget(QLabel("Тема:"))
        top_layout.addWidget(self.theme_combo)

        top_layout.addSpacing(20)

        self.log_path_label = QLabel(f" Папка с логами 📁 {SAVE_DIR}")
        self.log_path_label.setStyleSheet("font-size: 10px; color: #888;")
        top_layout.addWidget(self.log_path_label)

        log_folder_btn = QPushButton("📂")
        log_folder_btn.setFixedHeight(26)  # 🔥 Уменьшена высота кнопки
        log_folder_btn.clicked.connect(self.open_log_folder)
        top_layout.addWidget(log_folder_btn)

        # Кнопка изменения папки сохранения
        change_folder_btn = QPushButton("⚙️")
        change_folder_btn.setFixedHeight(26)
        change_folder_btn.clicked.connect(self.change_save_folder)
        top_layout.addWidget(change_folder_btn)

        # Кнопка очистки логов
        clear_logs_btn = QPushButton("🗑️")
        clear_logs_btn.setFixedHeight(26)
        clear_logs_btn.clicked.connect(self.clear_all_logs)
        top_layout.addWidget(clear_logs_btn)

        top_layout.addStretch()
        main_layout.addLayout(top_layout)

        # Список задач
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        self.tasks_container = QWidget()
        self.tasks_layout = QVBoxLayout(self.tasks_container)
        self.tasks_layout.addStretch()
        self.tasks_layout.setSpacing(5)
        scroll_area.setWidget(self.tasks_container)
        main_layout.addWidget(scroll_area)

        # Форма ввода
        input_form = QFrame()
        input_form.setFrameShape(QFrame.Shape.StyledPanel)
        form_layout = QVBoxLayout(input_form)

        self.task_input = QLineEdit()
        self.task_input.setPlaceholderText("Введите задачу (можно [текст](ссылка))")
        self.link_input = QLineEdit()
        self.link_input.setPlaceholderText("Ссылка (необязательно)")
        self.time_input = QLineEdit()
        self.time_input.setPlaceholderText("Время (минуты или 13:24-16:39). Оставьте пустым для запуска таймера.")

        form_layout.addWidget(QLabel("📝 Задача:"))
        form_layout.addWidget(self.task_input)
        form_layout.addWidget(QLabel("🔗 Ссылка:"))
        form_layout.addWidget(self.link_input)
        form_layout.addWidget(QLabel("⏱️ Время:"))
        form_layout.addWidget(self.time_input)

        # Кнопки внизу
        btn_layout = QHBoxLayout()
        add_btn = QPushButton("✅ Добавить задачу")
        add_btn.clicked.connect(self.add_task)
        stats_btn = QPushButton("📊 Статистика")
        stats_btn.clicked.connect(self.show_stats)
        save_btn = QPushButton("💾 Сохранить в Excel")
        save_btn.clicked.connect(self.save_excel_gui)
        quit_btn = QPushButton("🚪 Завершить день")
        quit_btn.clicked.connect(self.quit_app)

        btn_layout.addWidget(add_btn)
        btn_layout.addWidget(stats_btn)
        btn_layout.addWidget(save_btn)
        btn_layout.addWidget(quit_btn)

        form_layout.addLayout(btn_layout)
        main_layout.addWidget(input_form)

    def apply_theme(self):
        palette = self.palette()
        if current_theme == "dark":
            palette.setColor(QPalette.ColorRole.Window, QColor(30, 30, 30))
            palette.setColor(QPalette.ColorRole.WindowText, QColor(220, 220, 220))
            palette.setColor(QPalette.ColorRole.Base, QColor(40, 40, 40))
            palette.setColor(QPalette.ColorRole.AlternateBase, QColor(50, 50, 50))
            palette.setColor(QPalette.ColorRole.ToolTipBase, QColor(25, 25, 25))
            palette.setColor(QPalette.ColorRole.ToolTipText, QColor(220, 220, 220))
            palette.setColor(QPalette.ColorRole.Text, QColor(220, 220, 220))
            palette.setColor(QPalette.ColorRole.Button, QColor(50, 50, 50))
            palette.setColor(QPalette.ColorRole.ButtonText, QColor(220, 220, 220))
            palette.setColor(QPalette.ColorRole.BrightText, QColor(255, 0, 0))
            palette.setColor(QPalette.ColorRole.Link, QColor(42, 130, 218))
            palette.setColor(QPalette.ColorRole.Highlight, QColor(42, 130, 218))
            palette.setColor(QPalette.ColorRole.HighlightedText, QColor(30, 30, 30))
            self.setStyleSheet(f"""
                QMainWindow, QWidget {{ background-color: #1e1e1e; color: #dcdcdc; }}
                QLineEdit, QTextEdit {{ background-color: #2a2a2a; color: #dcdcdc; border: 1px solid #555; padding: 5px; }}
                QPushButton {{ 
                    background-color: #3a3a3a; 
                    color: #dcdcdc; 
                    border: 1px solid #555; 
                    padding: 4px 8px;
                    margin: 2px;
                    border-radius: 4px;
                }}
                QPushButton:hover {{ background-color: #4a4a4a; }}
                QScrollArea {{ border: none; }}
                QLabel {{ color: #dcdcdc; }}
                QComboBox {{ background-color: #2a2a2a; color: #dcdcdc; border: 1px solid #555; }}
            """)
            self.theme_combo.setCurrentIndex(0)
        else:
            palette.setColor(QPalette.ColorRole.Window, QColor(240, 240, 240))
            palette.setColor(QPalette.ColorRole.WindowText, QColor(0, 0, 0))
            palette.setColor(QPalette.ColorRole.Base, QColor(255, 255, 255))
            palette.setColor(QPalette.ColorRole.AlternateBase, QColor(245, 245, 245))
            palette.setColor(QPalette.ColorRole.ToolTipBase, QColor(255, 255, 220))
            palette.setColor(QPalette.ColorRole.ToolTipText, QColor(0, 0, 0))
            palette.setColor(QPalette.ColorRole.Text, QColor(0, 0, 0))
            palette.setColor(QPalette.ColorRole.Button, QColor(240, 240, 240))
            palette.setColor(QPalette.ColorRole.ButtonText, QColor(0, 0, 0))
            palette.setColor(QPalette.ColorRole.BrightText, QColor(255, 0, 0))
            palette.setColor(QPalette.ColorRole.Link, QColor(0, 0, 255))
            palette.setColor(QPalette.ColorRole.Highlight, QColor(0, 120, 215))
            palette.setColor(QPalette.ColorRole.HighlightedText, QColor(255, 255, 255))
            self.setStyleSheet("""
                QMainWindow, QWidget { background-color: #f0f0f0; color: #000; }
                QLineEdit, QTextEdit { background-color: #fff; color: #000; border: 1px solid #ccc; padding: 5px; }
                QPushButton { background-color: #f0f0f0; color: #000; border: 1px solid #ccc; padding: 4px 8px; margin: 2px; border-radius: 4px; }
                QPushButton:hover { background-color: #e0e0e0; }
                QScrollArea { border: none; }
                QLabel { color: #000; }
                QComboBox { background-color: #fff; color: #000; border: 1px solid #ccc; }
            """)
            self.theme_combo.setCurrentIndex(1)

        self.setPalette(palette)

    def toggle_theme(self):
        global current_theme
        current_theme = "light" if current_theme == "dark" else "dark"
        self.apply_theme()
        self.reload_all_task_widgets()

    def reload_all_task_widgets(self):
        for i in reversed(range(self.tasks_layout.count() - 1)):
            widget = self.tasks_layout.itemAt(i).widget()
            if widget:
                widget.deleteLater()
        self.load_tasks_to_ui()

    def load_tasks_to_ui(self):
        for i in reversed(range(self.tasks_layout.count() - 1)):
            widget = self.tasks_layout.itemAt(i).widget()
            if widget:
                widget.deleteLater()

        for idx, task in enumerate(tasks):
            task_widget = TaskItemWidget(task, idx, self)
            self.tasks_layout.insertWidget(self.tasks_layout.count() - 1, task_widget)
            self.tasks_widgets.append(task_widget)

    def add_task(self):
        task_text = self.task_input.text().strip()
        if not task_text:
            QMessageBox.warning(self, "Ошибка", "Введите задачу!")
            return

        is_completed = task_text.startswith("!")
        if is_completed:
            task_text = task_text[1:].strip()

        task_text, extracted_link = extract_markdown_links(task_text)
        link = self.link_input.text().strip() or extracted_link

        time_str = self.time_input.text().strip()

        if time_str:
            # Пользователь указал время — создаём ЗАВЕРШЁННУЮ задачу
            if "-" in time_str:
                minutes = parse_time_range(time_str)
            else:
                try:
                    minutes = float(time_str)
                except:
                    QMessageBox.warning(self, "Ошибка", "Неверный формат времени. Используйте число минут или формат 13:24-16:39")
                    return

            hours_hundredths = round(minutes / 60, 2)
            display_time_str = "<1 минуты" if minutes < 1 else f"{int(minutes)}:{int((minutes%1)*60):02d}"
            if "-" in time_str:
                display_time_str = time_str

            # ✅ ЗАВЕРШЁННАЯ задача — БЕЗ timer_start
            tasks.append({
                "id": str(uuid.uuid4()),
                "task": task_text,
                "link": link,
                "time_str": display_time_str,
                "minutes": round(minutes, 2),
                "hours_hundredths": hours_hundredths
            })

        else:
            # ✅ АКТИВНАЯ задача — с timer_start
            tasks.append({
                "id": str(uuid.uuid4()),
                "task": task_text,
                "link": link,
                "timer_start": time.time(),
                "minutes": 0.0,
                "hours_hundredths": 0.0,
                "time_str": "0:00"
            })

        save_backup()
        self.task_input.clear()
        self.link_input.clear()
        self.time_input.clear()
        self.load_tasks_to_ui()

    def edit_task(self, index):
        if index < 0 or index >= len(tasks):
            return

        task = tasks[index]
        dialog = EditTaskDialog(task, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            new_data = dialog.get_data()
            if new_data:
                tasks[index].update(new_data)
                save_backup()
                self.load_tasks_to_ui()

    def delete_task(self, index):
        if index < 0 or index >= len(tasks):
            return

        reply = QMessageBox.question(self, "Подтверждение", "Вы уверены, что хотите удалить эту задачу?",
                                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            tasks.pop(index)
            save_backup()
            self.load_tasks_to_ui()

    def show_stats(self):
        total_tasks = len(tasks)
        total_minutes = round(sum(t.get('minutes', 0) for t in tasks), 2)
        total_hours = round(sum(t.get('hours_hundredths', 0) for t in tasks), 2)

        link_minutes = {}
        for t in tasks:
            if t.get('link'):
                d = parse_domain(t['link'])
                link_minutes[d] = link_minutes.get(d, 0) + t.get('minutes', 0)
        top_links = sorted(link_minutes.items(), key=lambda x: x[1], reverse=True)[:3]

        task_minutes = {}
        for t in tasks:
            task_minutes[t['task']] = task_minutes.get(t['task'], 0) + t.get('minutes', 0)
        top_tasks = sorted(task_minutes.items(), key=lambda x: x[1], reverse=True)[:3]

        stats_text = f"📊 <b>Статистика</b>\n\n"
        stats_text += f"Всего задач: <b>{total_tasks}</b>\n"
        stats_text += f"Общее время: <b>{total_minutes} мин</b> / <b>{total_hours} ч</b>\n\n"

        if top_links:
            stats_text += "<b>Топ-3 ссылок по времени:</b>\n"
            for domain, mins in top_links:
                stats_text += f" • {domain}: {round(mins, 2)} мин\n"

        if top_tasks:
            stats_text += "\n<b>Топ-3 задач по времени:</b>\n"
            for name, mins in top_tasks:
                stats_text += f" • {name}: {round(mins, 2)} мин\n"

        QMessageBox.information(self, "Статистика", stats_text)

    def save_excel_gui(self):
        global SAVE_DIR
        options = QFileDialog.Option.DontUseNativeDialog
        folder = QFileDialog.getExistingDirectory(
            self,
            "Выберите папку для сохранения Excel-файла",
            SAVE_DIR,
            options=options
        )
        if folder:
            SAVE_DIR = folder
            self.log_path_label.setText(f"📁 {SAVE_DIR}")

        file_path = save_excel()
        if file_path:
            QMessageBox.information(self, "Успех", f"Файл сохранён:\n{file_path}")
            reply = QMessageBox.question(self, "Открыть папку?", "Открыть папку с файлом?",
                                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            if reply == QMessageBox.StandardButton.Yes:
                if sys.platform == "win32":
                    os.startfile(SAVE_DIR)
                elif sys.platform == "darwin":
                    os.system(f'open "{SAVE_DIR}"')
                else:
                    os.system(f'xdg-open "{SAVE_DIR}"')
        else:
            QMessageBox.critical(self, "Ошибка", "Не удалось сохранить Excel")

    def open_log_folder(self):
        if os.path.exists(LOG_DIR):
            if sys.platform == "win32":
                os.startfile(LOG_DIR)
            elif sys.platform == "darwin":
                os.system(f'open "{LOG_DIR}"')
            else:
                os.system(f'xdg-open "{LOG_DIR}"')
        else:
            QMessageBox.warning(self, "Папка не найдена", "Временная папка с логами ещё не создана.")


    def change_save_folder(self):
        """Изменить папку для сохранения Excel-файлов"""
        global SAVE_DIR
        options = QFileDialog.Option.DontUseNativeDialog
        folder = QFileDialog.getExistingDirectory(
            self,
            "Выберите папку для сохранения отчётов Excel",
            SAVE_DIR,
            options=options
        )
        if folder:
            SAVE_DIR = folder
            self.log_path_label.setText(f"📁 {SAVE_DIR}")
            QMessageBox.information(self, "Успех", f"Папка сохранения изменена на:\n{SAVE_DIR}")

    def clear_all_logs(self):
        """Очистить все логи (временные .txt) и Excel-файлы + очистить список задач"""
        reply = QMessageBox.question(
            self,
            "Подтверждение",
            "Вы уверены, что хотите удалить ВСЕ логи и сбросить список задач?\n"
            "Это действие необратимо.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply != QMessageBox.StandardButton.Yes:
            return

        try:
            # Удаляем все .txt в LOG_DIR
            if os.path.exists(LOG_DIR):
                for file in os.listdir(LOG_DIR):
                    if file.endswith(".txt"):
                        os.remove(os.path.join(LOG_DIR, file))

            # Удаляем все .xlsx в SAVE_DIR (опционально — можно спросить отдельно)
            if os.path.exists(SAVE_DIR):
                for file in os.listdir(SAVE_DIR):
                    if file.endswith(".xlsx") and file.startswith(datetime.date.today().strftime("%Y-%m-%d")):
                        os.remove(os.path.join(SAVE_DIR, file))

            # Очищаем список задач
            tasks.clear()
            save_backup()  # перезаписываем пустой backup
            self.load_tasks_to_ui()

            QMessageBox.information(self, "Очистка", "Все логи и задачи удалены.")

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось очистить логи:\n{str(e)}")


    def quit_app(self):
        save_excel()
        # Очищаем временный файл активных задач
        if os.path.exists(active_tasks_file):
            try:
                os.remove(active_tasks_file)
            except Exception as e:
                print(Fore.YELLOW + f"Не удалось удалить временный файл: {e}")
        QApplication.quit()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())