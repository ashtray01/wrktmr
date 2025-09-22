import os
import sys
import time
import datetime
import signal
import re
import tempfile
import ctypes
import uuid
import configparser
from urllib.parse import urlparse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter
from colorama import init, Fore, Style

# Инициализация colorama для цветного вывода в консоль
init(autoreset=True)

# Функция для установки темной темы для заголовка окна в Windows
def set_dark_title_bar_qt(window):
    """Устанавливает темную тему для заголовка окна в Windows."""
    if sys.platform != "win32":
        return
    try:
        hwnd = int(window.winId())
        dark_mode_attribute = 20  # Атрибут для темной темы
        value = ctypes.c_int(1)
        ctypes.windll.dwmapi.DwmSetWindowAttribute(
            hwnd,
            dark_mode_attribute,
            ctypes.byref(value),
            ctypes.sizeof(value)
        )
    except Exception as e:
        print(f"[Тёмный заголовок] Не удалось применить: {e}")

# Импорт необходимых модулей PyQt6
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                             QTextEdit, QLineEdit, QPushButton, QLabel, QScrollArea,
                             QFrame, QMessageBox, QFileDialog, QComboBox, QDialog,
                             QDialogButtonBox, QFormLayout, QSpinBox, QDoubleSpinBox,
                             QCheckBox, QGroupBox, QGridLayout)
from PyQt6.QtCore import Qt, QTimer, QUrl, QTime, QSize, QPoint
from PyQt6.QtGui import QFont, QDesktopServices, QColor, QTextCharFormat, QTextCursor, QPalette, QIcon

# Функция для получения пути к ресурсам (работает как в разработке, так и в собранном виде)
def resource_path(relative_path):
    """ Получает абсолютный путь к ресурсу, работает для dev и для PyInstaller """
    try:
        # PyInstaller создаёт временную папку и сохраняет путь в _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# Функции для получения директорий конфигурации и логов
def get_config_dir():
    """Получает путь к директории конфигурации."""
    temp_dir = tempfile.gettempdir()
    return os.path.join(temp_dir, "wrktmr-tmp")

def get_log_dir():
    """Получает путь к директории логов."""
    config_dir = get_config_dir()
    today = datetime.date.today()
    return os.path.join(config_dir, "logs", str(today.year), f"{today.month:02d}", f"{today.day:02d}")

# Путь к файлу настроек
CONFIG_FILE = os.path.join(get_config_dir(), "settings.ini")
_save_settings_lock = False

# Функции для загрузки и сохранения настроек
def load_settings():
    """Загружает настройки из файла."""
    config = configparser.ConfigParser(interpolation=None)
    if not os.path.exists(CONFIG_FILE):
        os.makedirs(os.path.dirname(CONFIG_FILE), exist_ok=True)
        # Устанавливаем новые параметры окна по умолчанию
        config['USER'] = {
            'save_dir': os.path.join(os.path.expanduser("~"), "Desktop"),
            'window_x': '14',
            'window_y': '357',
            'window_width': '579',
            'window_height': '698'
        }
        try:
            with open(CONFIG_FILE, 'w', encoding='utf-8') as configfile:
                config.write(configfile)
        except Exception as e:
            print(f"[ERROR] Не удалось создать файл настроек '{CONFIG_FILE}': {e}")
    else:
        try:
            config.read(CONFIG_FILE, encoding='utf-8')
            if 'USER' not in config:
                # Если секция USER отсутствует, создаем её с новыми параметрами по умолчанию
                config['USER'] = {
                    'save_dir': os.path.join(os.path.expanduser("~"), "Desktop"),
                    'window_x': '14',
                    'window_y': '357',
                    'window_width': '579',
                    'window_height': '698'
                }
                save_settings(config)
            else:
                # Проверяем наличие всех необходимых параметров окна, и если их нет - устанавливаем значения по умолчанию
                default_user_settings = {
                    'save_dir': os.path.join(os.path.expanduser("~"), "Desktop"),
                    'window_x': '14',
                    'window_y': '357',
                    'window_width': '579',
                    'window_height': '698'
                }
                for key, default_value in default_user_settings.items():
                    if key not in config['USER']:
                        config['USER'][key] = default_value
                save_settings(config)
        except Exception as e:
            print(f"[ERROR] Не удалось прочитать файл настроек '{CONFIG_FILE}': {e}")
            config['USER'] = {
                'save_dir': os.path.join(os.path.expanduser("~"), "Desktop"),
                'window_x': '14',
                'window_y': '357',
                'window_width': '579',
                'window_height': '698'
            }
    return config

def save_settings(config):
    """Сохраняет настройки в файл."""
    try:
        with open(CONFIG_FILE, 'w', encoding='utf-8') as configfile:
            config.write(configfile)
        #print(f"[DEBUG] Настройки успешно сохранены в '{CONFIG_FILE}'.")
    except Exception as e:
        print(f"[ERROR] Не удалось сохранить настройки в '{CONFIG_FILE}': {e}")

# Загрузка настроек при запуске
app_config = load_settings()
if 'USER' not in app_config:
    app_config['USER'] = {}
user_section = app_config['USER']

# Значения настроек по умолчанию
default_user_settings = {
    'save_dir': os.path.join(os.path.expanduser("~"), "Desktop"),
    'window_x': '14',
    'window_y': '357',
    'window_width': '579',
    'window_height': '698'
}

# Установка значений по умолчанию, если они отсутствуют
for key, default_value in default_user_settings.items():
    if key not in user_section:
        user_section[key] = default_value

# Сохранение настроек
save_settings(app_config)

# Получение директории логов и создание её, если она не существует
LOG_DIR = get_log_dir()
os.makedirs(LOG_DIR, exist_ok=True)

# Получение директории сохранения отчетов
SAVE_DIR = user_section.get('save_dir', os.path.join(os.path.expanduser("~"), "Desktop"))

# Глобальный список задач
tasks = []

# Путь к файлу активных задач
active_tasks_file = os.path.join(LOG_DIR, "active_tasks.tmp")

# Таймер для автоматического сохранения активных задач
active_save_timer = None

# Обработчик сигнала прерывания (Ctrl+C)
def signal_handler(sig, frame):
    """Обработчик сигнала прерывания (Ctrl+C)."""
    print(Fore.YELLOW + "\nОбнаружено Ctrl+C! Сохраняем данные...")
    save_backup()
    print(Fore.MAGENTA + "Данные сохранены. Программа завершена.")
    sys.exit(0)

# Регистрация обработчика сигнала
signal.signal(signal.SIGINT, signal_handler)

# Функция для извлечения домена из URL
def parse_domain(url):
    """Извлекает домен из URL."""
    try:
        parsed = urlparse(url)
        return parsed.netloc or url
    except:
        return url

# Функция для форматирования периода выполнения задачи
def format_time_period(start_timestamp, end_timestamp):
    """Форматирует период выполнения задачи как 'HH:MM-HH:MM'."""
    try:
        if start_timestamp is None or end_timestamp is None:
            return ""
        start_dt = datetime.datetime.fromtimestamp(start_timestamp)
        end_dt = datetime.datetime.fromtimestamp(end_timestamp)
        # Проверяем, перешли ли мы на следующий день
        if end_dt.date() > start_dt.date():
            # Можно добавить дату, но для простоты оставим только время
            # return f"{start_dt.strftime('%d.%m %H:%M')}-{end_dt.strftime('%d.%m %H:%M')}"
            # Или просто покажем время с пометкой "+1д"
            return f"{start_dt.strftime('%H:%M')}-{end_dt.strftime('%H:%M')} (+1д)"
        else:
            return f"{start_dt.strftime('%H:%M')}-{end_dt.strftime('%H:%M')}"
    except Exception as e:
        print(f"Ошибка форматирования периода: {e}")
        return ""

# Функция для форматирования времени начала и окончания задачи
def format_time_range(start_timestamp, end_timestamp=None, is_paused=False):
    """Форматирует время начала и окончания задачи."""
    try:
        if not isinstance(start_timestamp, (int, float)):
            start_str = "Начало: неизвестно"
        else:
            start_dt = datetime.datetime.fromtimestamp(start_timestamp)
            start_str = start_dt.strftime("%H:%M")
        if end_timestamp is None:
            if is_paused:
                return f"Начало {start_str} - На паузе"
            else:
                return f"Начало {start_str} - В процессе"
        else:
            end_dt = datetime.datetime.fromtimestamp(end_timestamp)
            end_str = end_dt.strftime("%H:%M")
            return f"Начало {start_str} - Завершена {end_str}"
    except Exception as e:
        print(f"Ошибка форматирования времени: {e}")
        return "Время неизвестно"

# Функция для получения ключа для группировки задач
def key_for_group(t):
    """Получает ключ для группировки задач."""
    if t.get("link"):
        return t["link"].strip().lower()
    return t["task"].strip().lower()

# Функция для сохранения резервной копии завершенных задач
def save_backup():
    """Сохраняет завершенные задачи в текстовый файл бэкапа."""
    today = datetime.date.today().strftime("%Y-%m-%d")
    log_file = os.path.join(LOG_DIR, f"{today}.txt")
    os.makedirs(os.path.dirname(log_file), exist_ok=True)
    # Открываем файл для записи (перезаписываем)
    with open(log_file, "w", encoding="utf-8") as f:
        for t in tasks:
            # Сохраняем только завершенные задачи (без активного таймера)
            if "timer_start" not in t:
                # Создаем копию, исключая ненужные ключи
                clean_t = {k: v for k, v in t.items() if k not in ['timer_start']}
                # Добавляем start и end timestamp, если они есть
                start_ts_str = str(clean_t.get('start_timestamp', ''))
                end_ts_str = str(clean_t.get('end_timestamp', ''))
                # Формат: task | link | time_str | hours_hundredths | start_timestamp | end_timestamp
                f.write(f"{clean_t['task']} | {clean_t['link']} | {clean_t['time_str']} | {clean_t['hours_hundredths']} | {start_ts_str} | {end_ts_str}\n")

# Функция для загрузки резервной копии завершенных задач
def load_backup():
    """Загружает завершенные задачи из текстового файла бэкапа."""
    today = datetime.date.today().strftime("%Y-%m-%d")
    log_file = os.path.join(LOG_DIR, f"{today}.txt")
    if os.path.exists(log_file):
        # Очищаем список задач перед загрузкой, кроме активных (на всякий случай, если load_backup вызывается отдельно)
        # Лучше этого избежать, но для корректной загрузки при старте это нужно учесть в порядке вызовов
        # Предполагаем, что активные задачи уже загружены
        initial_active_count = len([t for t in tasks if 'timer_start' in t])
        # Загружаем завершенные задачи
        with open(log_file, "r", encoding="utf-8") as f:
            for line in f:
                try:
                    parts = line.strip().split(" | ")
                    # Проверяем минимальное количество частей (без timestamp)
                    if len(parts) < 4:
                        continue
                    task, link, time_str, hours_hundredths_str = parts[0], parts[1], parts[2], parts[3]
                    # Обрабатываем время в минутах
                    if time_str == "<1 минуты":
                        minutes = 0.5
                    else:
                        if "-" in time_str:
                            minutes = parse_time_range(time_str)
                        else:
                            try:
                                mins, secs = map(int, time_str.split(":"))
                                minutes = mins + secs / 60
                            except:
                                minutes = float(time_str) if time_str.replace('.','',1).isdigit() else 0.0
                    # Создаем базовую структуру задачи
                    task_entry = {
                        "id": str(uuid.uuid4()),
                        "task": task,
                        "link": link,
                        "time_str": time_str,
                        "minutes": round(minutes, 2),
                        "hours_hundredths": float(hours_hundredths_str)
                    }
                    # Если в строке есть timestamp'ы, добавляем их
                    if len(parts) >= 6:
                        start_ts_str, end_ts_str = parts[4], parts[5]
                        if start_ts_str:
                            try:
                                task_entry["start_timestamp"] = float(start_ts_str)
                            except ValueError:
                                pass # Игнорируем ошибки преобразования
                        if end_ts_str:
                            try:
                                task_entry["end_timestamp"] = float(end_ts_str)
                            except ValueError:
                                pass # Игнорируем ошибки преобразования
                    tasks.append(task_entry)
                except Exception as e:
                    print(Fore.RED + f"Ошибка чтения backup: {e}")
        if tasks:
            # Подсчитываем только что загруженные завершенные задачи
            loaded_completed_count = len(tasks) - initial_active_count
            print(Fore.GREEN + f"Загружено {loaded_completed_count} задач из backup: {log_file}")
    # Логика удаления дубликатов активных задач оставлена в load_active_tasks

# Функция для сохранения активных задач во временный файл
def save_active_tasks():
    """Сохраняет активные задачи во временный файл.
    
    Формат строки в файле:
    id | task | link | timer_start | start_timestamp | paused_total | pause_history | is_paused
    где:
    - id: уникальный идентификатор задачи
    - task: текст задачи
    - link: ссылка задачи
    - timer_start: время запуска таймера (для вычисления текущего времени)
    - start_timestamp: метка времени начала задачи
    - paused_total: общее время в паузе (в секундах)
    - pause_history: список периодов паузы в формате "start1-end1,start2-end2,..."
    - is_paused: флаг, указывающий, находится ли задача на паузе (true/false)
    """
    try:
        active_tasks = [t for t in tasks if "timer_start" in t]
        if not active_tasks:
            if os.path.exists(active_tasks_file):
                os.remove(active_tasks_file)
            return
        # Открываем файл для записи (перезаписываем)
        with open(active_tasks_file, "w", encoding="utf-8") as f:
            for t in active_tasks:
                task_id = t.get("id", str(uuid.uuid4()))
                task_text = t['task']
                link = t['link']
                timer_start = t.get("timer_start", time.time())
                start_timestamp = t.get("start_timestamp", timer_start)
                paused_total = t.get("paused_total", 0.0)
                is_paused = t.get("is_paused", False)
                
                # Формируем строку истории пауз
                pause_history = ""
                if "pause_history" in t:
                    history_parts = []
                    for pause in t["pause_history"]:
                        start_pause = pause.get("start", "")
                        end_pause = pause.get("end", "")
                        history_parts.append(f"{start_pause}-{end_pause}")
                    pause_history = ",".join(history_parts)
                
                # Формат: id | task | link | timer_start | start_timestamp | paused_total | pause_history | is_paused
                f.write(f"{task_id} | {task_text} | {link} | {timer_start} | {start_timestamp} | {paused_total} | {pause_history} | {str(is_paused).lower()}\n")
    except Exception as e:
        print(Fore.RED + f"Ошибка сохранения активных задач: {e}")

# Функция для загрузки активных задач из временного файла
def load_active_tasks():
    """Загружает активные задачи из временного файла.
    
    При загрузке проверяется уникальность задач по id и по тексту/ссылке,
    чтобы избежать дублирования после перезапуска программы.
    Поддерживает формат с информацией о паузах.
    """
    if not os.path.exists(active_tasks_file):
        return
    try:
        # Перед загрузкой удаляем старые активные задачи из tasks, чтобы избежать дубликатов
        tasks[:] = [t for t in tasks if "timer_start" not in t]
        with open(active_tasks_file, "r", encoding="utf-8") as f:
            for line in f:
                try:
                    parts = line.strip().split(" | ")
                    if len(parts) < 4: # Минимум 4 части
                        continue
                    task_id, task, link, timer_start_str = parts[0], parts[1], parts[2], parts[3]
                    
                    # Проверка на дубликаты по id
                    if any(t.get("id") == task_id for t in tasks):
                        print(Fore.YELLOW + f"Пропущена дублирующаяся задача с id {task_id}")
                        continue
                    
                    # Проверка на дубликаты по тексту и ссылке
                    if any(t.get("task") == task and t.get("link") == link for t in tasks):
                        print(Fore.YELLOW + f"Пропущена дублирующаяся задача '{task}' с ссылкой '{link}'")
                        continue
                    
                    timer_start = float(timer_start_str)
                    
                    # Создаем активную задачу с базовыми значениями
                    active_task = {
                        "id": task_id,
                        "task": task,
                        "link": link,
                        "timer_start": timer_start,
                        "start_timestamp": timer_start,
                        "paused_total": 0.0,
                        "pause_history": [],
                        "is_paused": False,
                        "minutes": 0.0,
                        "hours_hundredths": 0.0,
                        "time_str": "0:00"
                    }
                    
                    # Если в строке есть дополнительные поля
                    if len(parts) >= 5:
                        try:
                            active_task["start_timestamp"] = float(parts[4])
                        except ValueError:
                            pass
                    
                    if len(parts) >= 6:
                        try:
                            active_task["paused_total"] = float(parts[5])
                        except ValueError:
                            pass
                    
                    if len(parts) >= 7:
                        pause_history_str = parts[6]
                        if pause_history_str:
                            try:
                                pause_history = []
                                for pause_part in pause_history_str.split(","):
                                    if "-" in pause_part:
                                        start_p, end_p = pause_part.split("-")
                                        pause_history.append({
                                            "start": float(start_p) if start_p and start_p.lower() not in ['none', ''] else None,
                                            "end": float(end_p) if end_p and end_p.lower() not in ['none', ''] else None
                                        })
                                active_task["pause_history"] = pause_history
                            except Exception:
                                active_task["pause_history"] = []
                    
                    if len(parts) >= 8:
                        is_paused_str = parts[7]
                        active_task["is_paused"] = is_paused_str.lower() == "true"
                        # Если задача на паузе, сохраняем накопленное время
                        if active_task["is_paused"]:
                            paused_total = active_task.get("paused_total", 0.0)
                            current_elapsed = time.time() - active_task["timer_start"]
                            active_task["paused_elapsed"] = current_elapsed - paused_total
                    
                    tasks.append(active_task)
                except Exception as e:
                    print(Fore.RED + f"Ошибка загрузки активной задачи: {e}")
        if tasks:
            active_count = len([t for t in tasks if 'timer_start' in t])
            print(Fore.GREEN + f"Загружено {active_count} активных задач из временного файла.")
    except Exception as e:
        print(Fore.RED + f"Ошибка чтения активных задач: {e}")

# Функция для удаления активной задачи из временного файла
def remove_active_task(index):
    """Удаляет активную задачу из временного файла."""
    try:
        if index < 0 or index >= len(tasks):
            return
        task_id = tasks[index].get("id")
        if not task_id:
            return
        if not os.path.exists(active_tasks_file):
            return
        remaining_lines = []
        removed = False
        with open(active_tasks_file, "r", encoding="utf-8") as f:
            for line in f:
                parts = line.strip().split(" | ")
                if len(parts) < 4:
                    remaining_lines.append(line.strip())
                    continue
                saved_id = parts[0]
                if saved_id != task_id:
                    remaining_lines.append(line.strip())
                else:
                    removed = True
        with open(active_tasks_file, "w", encoding="utf-8") as f:
            for line in remaining_lines:
                f.write(line + "\n")
        if not remaining_lines:
            os.remove(active_tasks_file)
        if removed:
            print(Fore.GREEN + f"Активная задача {task_id} удалена из временного файла.")
    except Exception as e:
        print(Fore.RED + f"Ошибка удаления активной задачи: {e}")

# Функция для группировки задач
def group_tasks():
    """Группирует задачи по ссылке или названию."""
    grouped = {}
    for t in tasks:
        link_key = t["link"].strip().lower() if t.get("link") else None
        if link_key:
            if link_key not in grouped:
                grouped[link_key] = {
                    "task": t["task"],
                    "link": t["link"],
                    "minutes": 0.0,
                    "hours_hundredths": 0.0,
                    "count": 0,
                    "key": link_key
                }
            grouped[link_key]["minutes"] += t["minutes"]
            grouped[link_key]["hours_hundredths"] += t["hours_hundredths"]
            grouped[link_key]["count"] += 1
        name_key = t["task"].strip().lower()
        if name_key not in grouped:
            grouped[name_key] = {
                "task": t["task"],
                "link": t["link"],
                "minutes": 0.0,
                "hours_hundredths": 0.0,
                "count": 0,
                "key": name_key
            }
        grouped[name_key]["minutes"] += t["minutes"]
        grouped[name_key]["hours_hundredths"] += t["hours_hundredths"]
        grouped[name_key]["count"] += 1
    return [g for g in grouped.values() if g["count"] > 1]

# Функция для парсинга временного диапазона
def parse_time_range(time_str):
    """Парсит временной диапазон."""
    try:
        start_str, end_str = time_str.split("-")
        h1, m1 = map(int, start_str.split(":"))
        h2, m2 = map(int, end_str.split(":"))
        start_minutes = h1 * 60 + m1
        end_minutes = h2 * 60 + m2
        if end_minutes < start_minutes:
            end_minutes += 24 * 60
        diff = end_minutes - start_minutes
        return diff
    except:
        return 0.0

# Функция для извлечения markdown-ссылок из текста
def extract_markdown_links(text):
    """Извлекает markdown-ссылки из текста."""
    pattern = r'\[([^\]]+)\]\(([^)]+)\)'
    matches = re.findall(pattern, text)
    if matches:
        display_text = matches[0][0]
        url = matches[0][1]
        clean_text = re.sub(pattern, display_text, text, count=1)
        return clean_text.strip(), url.strip()
    return text.strip(), ""

# Функция для стилизации Excel-файла
def style_tasks_and_summary(ws, n_tasks_rows, grouped, group_key_to_rows):
    """Стилизует Excel-файл."""
    thin = Border(left=Side(style='thin'),
                  right=Side(style='thin'),
                  top=Side(style='thin'),
                  bottom=Side(style='thin'))
    palette = [
        "FFF2CC", "E2EFDA", "DDEBF7", "FCE4D6", "EDEDED",
        "CCE5FF", "FFD966", "C6E0B4", "D9E1F2", "FFEB9C",
        "E7E6E6", "D5E8D4", "D0CECE", "F8CBAD", "C9DAF8"
    ]
    def color_for_key(k):
        if not k:
            return None
        idx = abs(hash(k)) % len(palette)
        return PatternFill(start_color=palette[idx], end_color=palette[idx], fill_type="solid")
    max_row = ws.max_row
    max_col = ws.max_column
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for col in range(1, max_col + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(vertical="center", horizontal="left")
        c.fill = header_fill
        c.border = thin
    for r in range(2, 2 + n_tasks_rows - 1 if n_tasks_rows > 0 else 1):
        zebra = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid") if (r % 2 == 0) else None
        if r in group_key_to_rows["row_to_key"]:
            k = group_key_to_rows["row_to_key"][r]
            fill = color_for_key(k)
        else:
            fill = zebra
        for col in range(1, max_col + 1):
            cell = ws.cell(row=r, column=col)
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin
    total_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "ИТОГО":
            total_row = r
            break
    if total_row:
        for col in range(1, max_col + 1):
            c = ws.cell(row=total_row, column=col)
            c.font = Font(bold=True)
            c.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            c.border = thin
    summary_header_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "СВОДКА (объединено по совпадению ссылки или названия)":
            summary_header_row = r
            break
    if summary_header_row:
        for col in range(1, max_col + 1):
            c = ws.cell(row=summary_header_row, column=col)
            c.font = Font(bold=True)
            c.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            c.border = thin
        sr = summary_header_row + 1
        while sr <= ws.max_row:
            a = ws.cell(row=sr, column=1).value
            b = ws.cell(row=sr, column=2).value
            c3 = ws.cell(row=sr, column=3).value
            if a is None and b is None and c3 is None:
                break
            matched_key = None
            for g in grouped:
                if a and a == g["task"]:
                    matched_key = g["key"]
                    break
                if b and b == parse_domain(g["link"] or ""):
                    matched_key = g["key"]
                    break
            if matched_key:
                fill = color_for_key(matched_key)
                if fill:
                    for col in range(1, max_col + 1):
                        cell = ws.cell(row=sr, column=col)
                        cell.fill = fill
                        cell.border = thin
            sr += 1
    if n_tasks_rows > 0:
        ws.auto_filter.ref = f"A1:D{1 + n_tasks_rows}"
    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=max_col):
        for cell in row:
            if not cell.border:
                cell.border = thin
            if not cell.alignment:
                cell.alignment = Alignment(vertical="center")
    for col in range(1, max_col + 1):
        column = get_column_letter(col)
        max_len = 0
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[column].width = max(10, min(60, max_len + 2))

# Функция для сохранения данных в Excel
def save_excel():
    """Сохраняет данные в Excel-файл."""
    global SAVE_DIR
    today = datetime.date.today().strftime("%Y-%m-%d")
    file_name = os.path.join(SAVE_DIR, f"{today}.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tasks"
    # Добавляем новые заголовки
    ws.append(["Задача", "Ссылка (домен)", "Время (мин:сек)", "Время (часы в сотых)", "Занято, мин", "Период выполнения"])
    total_minutes = 0.0
    total_hours_hundredths = 0.0
    total_occupied_minutes = 0.0 # Для подсчета итога по "Занято, мин"
    keys = [key_for_group(t) for t in tasks]
    dup_keys = {k for k, c in Counter(keys).items() if c > 1}
    row_to_key = {}
    for t in tasks:
        domain = parse_domain(t['link']) if t['link'] else ""
        row_index = ws.max_row + 1
        ws.cell(row=row_index, column=1, value=t['task'])
        if t['link']:
            link_cell = ws.cell(row=row_index, column=2, value=domain)
            link_cell.hyperlink = t['link']
            link_cell.style = "Hyperlink"
        else:
            ws.cell(row=row_index, column=2, value="")
        ws.cell(row=row_index, column=3, value=t['time_str'])
        ws.cell(row=row_index, column=4, value=t['hours_hundredths'])
        # Добавляем новые колонки
        # "Занято, мин" - это просто minutes, округленные до 2 знаков
        occupied_minutes = round(t.get('minutes', 0), 2)
        ws.cell(row=row_index, column=5, value=occupied_minutes)
        # "Период выполнения"
        period_str = format_time_period(t.get('start_timestamp'), t.get('end_timestamp'))
        ws.cell(row=row_index, column=6, value=period_str)
        k = key_for_group(t)
        if k in dup_keys:
            row_to_key[row_index] = k
        total_minutes += t['minutes']
        total_hours_hundredths += t['hours_hundredths']
        total_occupied_minutes += occupied_minutes # Суммируем для итога
    n_tasks_rows = len(tasks) + 1
    # Обновляем строку итогов
    ws.append([])
    ws.append(["ИТОГО", "", f"{round(total_minutes, 2)} мин", f"{round(total_hours_hundredths, 2)} ч", f"{round(total_occupied_minutes, 2)} мин", ""])
    grouped = group_tasks()
    if grouped:
        ws.append([])
        ws.append(["СВОДКА (объединено по совпадению ссылки или названия)"])
        for gt in grouped:
            ws.append([
                gt["task"],
                parse_domain(gt["link"]) if gt["link"] else "",
                f"{round(gt['minutes'], 2)} мин",
                f"{round(gt['hours_hundredths'], 2)} ч",
                "", # Пусто для сводки
                ""  # Пусто для сводки
            ])
    # Передаем n_tasks_rows+1, так как добавили 2 колонки
    style_tasks_and_summary(
        ws=ws,
        n_tasks_rows=n_tasks_rows,
        grouped=grouped,
        group_key_to_rows={"row_to_key": row_to_key}
    )
    try:
        wb.save(file_name)
        print(Fore.GREEN + f"Excel-отчёт сохранён: {file_name}")
        return file_name
    except Exception as e:
        print(Fore.RED + f"Ошибка при сохранении Excel: {e}")
        return None

# Виджет для отображения задачи
class TaskItemWidget(QFrame):
    """Виджет для отображения задачи."""
    def __init__(self, task_data, index, parent=None):
        super().__init__(parent)
        self.task_data = task_data
        self.index = index
        self.parent_window = parent
        self.timer_label = None
        self.timer = None
        # Флаг для отображения подтверждения завершения
        self.confirmation_visible = False
        self.setup_ui()
        if "timer_start" in self.task_data:
            self.start_live_timer()

    def setup_ui(self):
        """Настраивает пользовательский интерфейс виджета задачи."""
        self.setStyleSheet("""
                    TaskItemWidget {
                        background-color: #2a2a2a;
                        border-radius: 6px;
                        border: 1px solid #444;
                        margin: 4px;
                    }
                """)
        self.setFrameShadow(QFrame.Shadow.Raised)
        main_layout = QHBoxLayout()
        main_layout.setContentsMargins(2, 2, 2, 2)
        main_layout.setSpacing(5)
        left_layout = QVBoxLayout()
        left_layout.setSpacing(3)
        task_label = QLabel()
        task_text = self.task_data['task']
        task_label.setTextFormat(Qt.TextFormat.RichText)
        task_label.setText(self.format_markdown_links(task_text))
        task_label.setWordWrap(True)
        task_label.setTextInteractionFlags(Qt.TextInteractionFlag.TextBrowserInteraction)
        task_label.linkActivated.connect(self.open_link)
        task_label.setStyleSheet("font-size: 13px;")
        left_layout.addWidget(task_label)
        if self.task_data['link']:
            domain = parse_domain(self.task_data['link'])
            link_label = QLabel(f"<a href='{self.task_data['link']}' style='color: #64b5f6; font-size: 11px;'>🌐 {domain}</a>")
            link_label.setTextFormat(Qt.TextFormat.RichText)
            link_label.setTextInteractionFlags(Qt.TextInteractionFlag.TextBrowserInteraction)
            link_label.linkActivated.connect(self.open_link)
            left_layout.addWidget(link_label)
        self.time_display = QLabel()
        self.update_time_display()
        self.time_display.setStyleSheet("font-size: 11px; color: #aaa;")
        left_layout.addWidget(self.time_display)
        main_layout.addLayout(left_layout)
        button_layout = QVBoxLayout()
        button_layout.setSpacing(5)
        button_layout.addStretch()
        button_width = int(self.parent_window.width() * 0.2) if self.parent_window else 150
        button_height = 32
        if "timer_start" in self.task_data:
            # Кнопка паузы/возобновления
            self.pause_resume_btn = QPushButton("Пауза")
            self.pause_resume_btn.setFixedSize(button_width, button_height)
            self.pause_resume_btn.clicked.connect(self.toggle_pause)
            # Устанавливаем стиль в зависимости от состояния паузы при инициализации
            if self.task_data.get("is_paused", False):
                self.pause_resume_btn.setText("Продолжить")
                self.pause_resume_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #2e7d32; /* Темно-зеленый */
                        color: white;
                        border: 1px solid #1b5e20;
                        border-radius: 4px;
                        padding: 4px 8px;
                    }
                    QPushButton:hover {
                        background-color: #388e3c;
                    }
                """)
            else:
                self.pause_resume_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #7f6000; /* Темно-желтый */
                        color: white;
                        border: 1px solid #5d4000;
                        border-radius: 4px;
                        padding: 4px 8px;
                    }
                    QPushButton:hover {
                        background-color: #997a00;
                    }
                """)
            button_layout.addWidget(self.pause_resume_btn)
            
            # Кнопка завершения
            self.finish_btn = QPushButton("Завершить")
            self.finish_btn.setFixedSize(button_width, button_height)
            self.finish_btn.clicked.connect(self.show_confirmation)
            self.finish_btn.setStyleSheet("""
                QPushButton {
                    background-color: #660000; /* Темно-красный */
                    color: white;
                    border: 1px solid #550000;
                    border-radius: 4px;
                    padding: 4px 8px;
                }
                QPushButton:hover {
                    background-color: #A60000;
                }
            """)
            button_layout.addWidget(self.finish_btn)
            
            # Кнопки подтверждения завершения (изначально скрыты)
            self.confirm_yes_btn = QPushButton("Да")
            self.confirm_yes_btn.setFixedSize(button_width, button_height)
            self.confirm_yes_btn.clicked.connect(self.confirm_finish_task)
            self.confirm_yes_btn.setStyleSheet("""
                QPushButton {
                    background-color: #388e3c; /* Зеленый */
                    color: white;
                    border: 1px solid #2e7d32;
                    border-radius: 4px;
                    padding: 4px 8px;
                }
                QPushButton:hover {
                    background-color: #4caf50;
                }
            """)
            self.confirm_no_btn = QPushButton("Нет")
            self.confirm_no_btn.setFixedSize(button_width, button_height)
            self.confirm_no_btn.clicked.connect(self.hide_confirmation)
            self.confirm_no_btn.setStyleSheet("""
                QPushButton {
                    background-color: #f44336; /* Красный */
                    color: white;
                    border: 1px solid #d32f2f;
                    border-radius: 4px;
                    padding: 4px 8px;
                }
                QPushButton:hover {
                    background-color: #ef5350;
                }
            """)
            self.confirm_yes_btn.setVisible(False)
            self.confirm_no_btn.setVisible(False)
            button_layout.addWidget(self.confirm_yes_btn)
            button_layout.addWidget(self.confirm_no_btn)
        else:
            # Кнопки для завершенных задач
            edit_btn = QPushButton("✏️ Изменить")
            edit_btn.setFixedSize(button_width, button_height)
            edit_btn.clicked.connect(self.edit_task)
            delete_btn = QPushButton("🗑️ Удалить")
            delete_btn.setFixedSize(button_width, button_height)
            delete_btn.clicked.connect(self.delete_task)
            button_layout.addWidget(edit_btn)
            button_layout.addWidget(delete_btn)
        button_layout.addStretch()
        main_layout.addLayout(button_layout)
        self.setLayout(main_layout)
        self.setFixedHeight(90) # Устанавливаем высоту элемента

    def format_markdown_links(self, text):
        """Форматирует markdown-ссылки в HTML."""
        pattern = r'\[([^\]]+)\]\(([^)]+)\)'
        def repl(match):
            display = match.group(1)
            url = match.group(2)
            return f'<a href="{url}" style="color: #64b5f6;">{display}</a>'
        return re.sub(pattern, repl, text)

    def open_link(self, url):
        """Открывает ссылку в браузере."""
        QDesktopServices.openUrl(QUrl(url))

    def update_time_display(self):
        """Обновляет отображение времени задачи."""
        if "timer_start" in self.task_data:
            # Вычисляем общее время в паузе
            paused_total = self.task_data.get("paused_total", 0.0)
            # Если задача на паузе, отображаем накопленное время
            if self.task_data.get("is_paused", False):
                elapsed = self.task_data.get("paused_elapsed", 0.0)
                time_range_str = format_time_range(self.task_data['start_timestamp'], is_paused=True)
                self.time_display.setStyleSheet("font-size: 11px; color: #ffd600;") # Желтый цвет при паузе
            else:
                # Если задача активна, вычисляем текущее время
                elapsed = time.time() - self.task_data['timer_start'] - paused_total
                time_range_str = format_time_range(self.task_data['start_timestamp'])
                self.time_display.setStyleSheet("font-size: 11px; color: #69f0ae;") # Зеленый цвет при активности

            minutes = int(elapsed // 60)
            hours_hundredths = round(elapsed / 3600, 2)
            anim_chars = ['|', '/', '-', '\\']
            if not hasattr(self, '_anim_index'):
                self._anim_index = 0
            else:
                self._anim_index = (self._anim_index + 1) % 4
            anim_char = anim_chars[self._anim_index]
            self.time_display.setText(f"{minutes} мин ({hours_hundredths} ч) {time_range_str} {anim_char}")
        else:
            minutes = int(self.task_data['minutes'])
            hours_hundredths = self.task_data['hours_hundredths']
            start_ts = self.task_data.get('start_timestamp')
            end_ts = self.task_data.get('end_timestamp')
            time_range_str = format_time_range(start_ts, end_ts) if start_ts else ""
            self.time_display.setStyleSheet("font-size: 11px; color: #aaa;") # Серый цвет для завершенных задач
            self.time_display.setText(f"{minutes} мин ({hours_hundredths} ч) {time_range_str}")

    def start_live_timer(self):
        """Запускает таймер для обновления времени в реальном времени."""
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.update_time_display)
        self.timer.start(1000)

    def toggle_pause(self):
        """Переключает состояние паузы задачи."""
        if "timer_start" not in self.task_data:
            return

        is_paused = self.task_data.get("is_paused", False)

        if is_paused:
            # Возобновляем задачу
            self.task_data["is_paused"] = False
            # Завершаем текущий период паузы
            if "pause_history" in self.task_data and self.task_data["pause_history"]:
                last_pause = self.task_data["pause_history"][-1]
                if last_pause.get("start") and last_pause.get("end") is None:
                    last_pause["end"] = time.time()
            # Обновляем время запуска таймера с учетом паузы
            self.task_data["timer_start"] = time.time() - (time.time() - self.task_data["timer_start"] - self.task_data.get("paused_total", 0.0))
            self.pause_resume_btn.setText("Пауза")
            self.pause_resume_btn.setStyleSheet("""
                QPushButton {
                    background-color: #7f6000; /* Темно-желтый */
                    color: white;
                    border: 1px solid #5d4000;
                    border-radius: 4px;
                    padding: 4px 8px;
                }
                QPushButton:hover {
                    background-color: #997a00;
                }
            """)
        else:
            # Ставим задачу на паузу
            self.task_data["is_paused"] = True
            # Начинаем новый период паузы
            if "pause_history" not in self.task_data:
                self.task_data["pause_history"] = []
            self.task_data["pause_history"].append({
                "start": time.time(),
                "end": None
            })
            # Сохраняем накопленное время на момент паузы
            paused_total = self.task_data.get("paused_total", 0.0)
            current_elapsed = time.time() - self.task_data["timer_start"]
            self.task_data["paused_elapsed"] = current_elapsed - paused_total
            self.pause_resume_btn.setText("Продолжить")
            self.pause_resume_btn.setStyleSheet("""
                QPushButton {
                    background-color: #2e7d32; /* Темно-зеленый */
                    color: white;
                    border: 1px solid #1b5e20;
                    border-radius: 4px;
                    padding: 4px 8px;
                }
                QPushButton:hover {
                    background-color: #388e3c;
                }
            """)

        # Сохраняем изменения
        save_active_tasks()
        self.update_time_display()

    def show_confirmation(self):
        """Показывает кнопки подтверждения завершения задачи."""
        self.confirmation_visible = True
        self.finish_btn.setVisible(False)
        self.pause_resume_btn.setVisible(False)
        self.confirm_yes_btn.setVisible(True)
        self.confirm_no_btn.setVisible(True)

    def hide_confirmation(self):
        """Скрывает кнопки подтверждения завершения задачи."""
        self.confirmation_visible = False
        self.finish_btn.setVisible(True)
        self.pause_resume_btn.setVisible(True)
        self.confirm_yes_btn.setVisible(False)
        self.confirm_no_btn.setVisible(False)

    def confirm_finish_task(self):
        """Подтверждает завершение задачи."""
        self.finish_task()

    def finish_task(self):
        """Завершает задачу и сохраняет её данные."""
        if "timer_start" not in self.task_data:
            return
        remove_active_task(self.index)
        
        # Вычисляем общее время выполнения задачи
        timer_start = self.task_data['timer_start']
        paused_total = self.task_data.get("paused_total", 0.0)
        # Завершаем все открытые периоды паузы
        if "pause_history" in self.task_data:
            for pause in self.task_data["pause_history"]:
                if pause.get("start") and pause.get("end") is None:
                    pause["end"] = time.time()
            # Пересчитываем общее время паузы
            paused_total = 0.0
            for pause in self.task_data["pause_history"]:
                start_p = pause.get("start")
                end_p = pause.get("end")
                if start_p and end_p:
                    paused_total += (end_p - start_p)
        
        # Вычисляем активное время (без учета пауз)
        elapsed = time.time() - timer_start - paused_total
        minutes = elapsed / 60
        hours_hundredths = round(minutes / 60, 2)
        time_str = "<1 минуты" if minutes < 1 else f"{int(minutes)}:{int((minutes%1)*60):02d}"
        tasks[self.index].update({
            "minutes": round(minutes, 2),
            "hours_hundredths": hours_hundredths,
            "time_str": time_str,
            "end_timestamp": time.time()
        })
        del tasks[self.index]["timer_start"]
        if "is_paused" in tasks[self.index]:
            del tasks[self.index]["is_paused"]
        if "paused_total" in tasks[self.index]:
            del tasks[self.index]["paused_total"]
        if "pause_history" in tasks[self.index]:
            del tasks[self.index]["pause_history"]
        if self.timer:
            self.timer.stop()
        save_backup()
        self.parent_window.load_tasks_to_ui()

    def edit_task(self):
        """Открывает диалог редактирования задачи."""
        self.parent_window.edit_task(self.index)

    def delete_task(self):
        """Удаляет задачу."""
        self.parent_window.delete_task(self.index)

# Функция для завершения активной задачи
def finish_active_task(task_index):
    """Преобразует активную задачу в завершенную, фиксируя текущее время как окончание."""
    if task_index < 0 or task_index >= len(tasks):
        return None
    task = tasks[task_index]
    # Проверяем, является ли задача активной
    if "timer_start" not in task:
        return None
    # Вычисляем время выполнения
    timer_start = task['timer_start']
    paused_total = task.get("paused_total", 0.0)
    # Завершаем все открытые периоды паузы
    if "pause_history" in task:
        for pause in task["pause_history"]:
            if pause.get("start") and pause.get("end") is None:
                pause["end"] = time.time()
        # Пересчитываем общее время паузы
        paused_total = 0.0
        for pause in task["pause_history"]:
            start_p = pause.get("start")
            end_p = pause.get("end")
            if start_p and end_p:
                paused_total += (end_p - start_p)
    
    # Вычисляем активное время (без учета пауз)
    elapsed = time.time() - timer_start - paused_total
    minutes = elapsed / 60
    hours_hundredths = round(minutes / 60, 2)
    # Форматируем строку времени
    time_str = "<1 минуты" if minutes < 1 else f"{int(minutes)}:{int((minutes%1)*60):02d}"
    # Создаем копию задачи и обновляем её данные
    finished_task = task.copy()
    finished_task.update({
        "minutes": round(minutes, 2),
        "hours_hundredths": hours_hundredths,
        "time_str": time_str,
        "end_timestamp": time.time()  # Текущее время как время завершения
    })
    # Удаляем таймер-специфичные поля
    finished_task.pop("timer_start", None)
    finished_task.pop("is_paused", None)
    finished_task.pop("paused_total", None)
    finished_task.pop("pause_history", None)
    return finished_task

# Диалог для редактирования задачи
class EditTaskDialog(QDialog):
    """Диалог для редактирования задачи."""
    def __init__(self, task_data, parent=None):
        super().__init__(parent)
        self.task_data = task_data
        self.setWindowTitle("✏️ Редактировать задачу")
        self.resize(500, 200)
        self.setup_ui()

    def setup_ui(self):
        """Настраивает пользовательский интерфейс диалога редактирования."""
        layout = QFormLayout()
        self.task_input = QLineEdit(self.task_data.get("task", ""))
        self.link_input = QLineEdit(self.task_data.get("link", ""))
        minutes_value = self.task_data.get("minutes", 0.0)
        self.time_input = QLineEdit(str(minutes_value))
        layout.addRow("📝 Задача:", self.task_input)
        layout.addRow("🔗 Ссылка:", self.link_input)
        layout.addRow("⏱️ Время (в минутах):", self.time_input)
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)
        self.setLayout(layout)

    def get_data(self):
        """Получает данные из диалога редактирования."""
        task_text = self.task_input.text().strip()
        link = self.link_input.text().strip()
        time_input_text = self.time_input.text().strip()
        # Инициализируем timestamp'ы
        start_timestamp = time.time()
        end_timestamp = start_timestamp
        try:
            # Проверяем, является ли ввод диапазоном времени
            if "-" in time_input_text:
                # Проверяем формат диапазона времени
                try:
                    start_str, end_str = time_input_text.split("-")
                    if len(start_str.split(":")) != 2 or len(end_str.split(":")) != 2:
                        raise ValueError("Неверный формат диапазона времени")
                    h1, m1 = map(int, start_str.split(":"))
                    h2, m2 = map(int, end_str.split(":"))
                    # Проверяем корректность времени
                    if not (0 <= h1 <= 23 and 0 <= m1 <= 59 and 0 <= h2 <= 23 and 0 <= m2 <= 59):
                        raise ValueError("Некорректное время")
                    today = datetime.date.today()
                    start_dt = datetime.datetime(today.year, today.month, today.day, h1, m1)
                    end_dt = datetime.datetime(today.year, today.month, today.day, h2, m2)
                    if end_dt < start_dt:
                        end_dt += datetime.timedelta(days=1)
                    # Вычисляем разницу в минутах
                    diff_seconds = (end_dt - start_dt).total_seconds()
                    minutes = diff_seconds / 60
                    # Проверяем адекватность продолжительности
                    if minutes > 24 * 60:
                        raise ValueError("Продолжительность задачи не может превышать 24 часа")
                    start_timestamp = start_dt.timestamp()
                    end_timestamp = end_dt.timestamp()
                except Exception as e:
                    QMessageBox.warning(self, "Ошибка", f"Неверный формат диапазона времени или время вне допустимых пределов: {str(e)}\nИспользуйте формат HH:MM-HH:MM (например, 13:24-16:39)")
                    return None
            else:
                # Обычное число минут
                minutes = float(time_input_text)
                # Проверяем адекватность продолжительности
                if minutes <= 0:
                    raise ValueError("Время должно быть положительным")
                if minutes > 24 * 60:  # Более 24 часов
                    raise ValueError("Время не может превышать 24 часа (1440 минут)")
                start_timestamp = time.time()
                end_timestamp = start_timestamp + (minutes * 60)
        except ValueError as e:
            QMessageBox.warning(self, "Ошибка", f"Неверный формат времени: {str(e)}\nВведите положительное число минут (не более 1440) или диапазон времени в формате HH:MM-HH:MM")
            return None
        hours_hundredths = round(minutes / 60, 2)
        time_str = "<1 минуты" if minutes < 1 else f"{int(minutes)}:{int((minutes%1)*60):02d}"
        if "-" in time_input_text:
            time_str = time_input_text
        return {
            "task": task_text,
            "link": link,
            "time_str": time_str,
            "minutes": round(minutes, 2),
            "hours_hundredths": hours_hundredths,
            "start_timestamp": start_timestamp,
            "end_timestamp": end_timestamp
        }

# Диалог настроек
class SettingsDialog(QDialog):
    """Диалог настроек приложения."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("⚙️ Настройки")
        self.resize(500, 300)
        self.setup_ui()

    def setup_ui(self):
        """Настраивает пользовательский интерфейс диалога настроек."""
        layout = QVBoxLayout()
        
        # Группа настроек путей
        paths_group = QGroupBox("Пути")
        paths_layout = QFormLayout()
        
        self.save_dir_input = QLineEdit(SAVE_DIR)  # Используем глобальную переменную SAVE_DIR
        self.save_dir_btn = QPushButton("Обзор...")
        self.save_dir_btn.clicked.connect(self.browse_save_dir)
        save_dir_layout = QHBoxLayout()
        save_dir_layout.addWidget(self.save_dir_input)
        save_dir_layout.addWidget(self.save_dir_btn)
        paths_layout.addRow("Папка для отчетов:", save_dir_layout)
        
        self.log_dir_input = QLineEdit(LOG_DIR)  # Используем глобальную переменную LOG_DIR
        self.log_dir_btn = QPushButton("Обзор...")
        self.log_dir_btn.clicked.connect(self.browse_log_dir)
        log_dir_layout = QHBoxLayout()
        log_dir_layout.addWidget(self.log_dir_input)
        log_dir_layout.addWidget(self.log_dir_btn)
        paths_layout.addRow("Папка для логов:", log_dir_layout)
        
        paths_group.setLayout(paths_layout)
        layout.addWidget(paths_group)
        
        # Кнопки OK/Cancel
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
        
        self.setLayout(layout)

    def browse_save_dir(self):
        """Открывает диалог выбора папки для сохранения отчетов."""
        folder = QFileDialog.getExistingDirectory(self, "Выберите папку для сохранения отчётов", self.save_dir_input.text())
        if folder:
            self.save_dir_input.setText(folder)

    def browse_log_dir(self):
        """Открывает диалог выбора папки для логов."""
        folder = QFileDialog.getExistingDirectory(self, "Выберите папку для логов", self.log_dir_input.text())
        if folder:
            self.log_dir_input.setText(folder)

    def get_data(self):
        """Получает данные из диалога настроек."""
        return {
            "save_dir": self.save_dir_input.text(),
            "log_dir": self.log_dir_input.text()
        }

# Главное окно приложения
class MainWindow(QMainWindow):
    """Главное окно приложения."""
    def __init__(self):
        super().__init__()
        x = int(app_config['USER'].get('window_x', 100))
        y = int(app_config['USER'].get('window_y', 100)) - 31  # Костыль: коррекция позиции по Y
        width = int(app_config['USER'].get('window_width', 900))
        height = int(app_config['USER'].get('window_height', 700))
        self.setGeometry(x, y, width, height)
        self.setWindowTitle("wrktmr v034.5")
        icon_path = resource_path(os.path.join('icon', 'icon.ico'))
        if os.path.exists(icon_path):
            icon = QIcon(icon_path)
            self.setWindowIcon(icon)
            QApplication.instance().setWindowIcon(icon) 
        else:
            print(f"[DEBUG] Иконка не найдена по пути: {icon_path}")
        self.tasks_widgets = []
        load_active_tasks()
        load_backup()
        self.init_ui()
        self.apply_theme()
        self.load_tasks_to_ui()
        self.start_active_save_timer()
        self.show()
        self.destroyed.connect(self.on_window_destroyed)
        set_dark_title_bar_qt(self)
        self.update_datetime()

    def update_datetime(self):
        """Обновляет отображение текущей даты и времени."""
        now = datetime.datetime.now()
        days = {
            0: "Понедельник",
            1: "Вторник",
            2: "Среда",
            3: "Четверг",
            4: "Пятница",
            5: "Суббота",
            6: "Воскресенье"
        }
        day_name = days[now.weekday()]
        self.datetime_label.setText(now.strftime("%d.%m.%Y %H:%M:%S") + " " + day_name)
        QTimer.singleShot(1000, self.update_datetime)

    # def on_window_move_or_resize(self, event):
    #     """Обработчик перемещения или изменения размера окна."""
    #     app_config['USER']['window_x'] = str(self.x())
    #     app_config['USER']['window_y'] = str(self.y())
    #     app_config['USER']['window_width'] = str(self.width())
    #     app_config['USER']['window_height'] = str(self.height())
    #     save_settings(app_config)
    #     if isinstance(event, PyQt6.QtGui.QMoveEvent):
    #         super().moveEvent(event)
    #     else:
    #         super().resizeEvent(event)

    def moveEvent(self, event):
        """Обработчик перемещения окна."""
        self.save_window_state()
        super().moveEvent(event)

    def resizeEvent(self, event):
        """Обработчик изменения размера окна."""
        self.save_window_state()
        super().resizeEvent(event)

    def save_window_state(self):
        """Сохраняет состояние окна (позицию и размер)."""
        try:
            app_config['USER']['window_x'] = str(self.x())
            app_config['USER']['window_y'] = str(self.y())
            app_config['USER']['window_width'] = str(self.width())
            app_config['USER']['window_height'] = str(self.height())
            save_settings(app_config)
        except Exception as e:
            print(f"[ERROR_save_window_state] {e}")
            
    def start_active_save_timer(self):
        """Запускает таймер для автоматического сохранения активных задач."""
        global active_save_timer
        if active_save_timer:
            return
        def save_loop():
            save_active_tasks()
            active_save_timer = QTimer.singleShot(10000, save_loop)
        save_loop()

    def init_ui(self):
        """Инициализирует пользовательский интерфейс главного окна."""
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        top_layout = QHBoxLayout()
        self.datetime_label = QLabel()
        self.datetime_label.setStyleSheet("font-size: 12px; color: #888;")
        top_layout.addWidget(self.datetime_label)
        top_layout.addStretch()
        log_folder_btn = QPushButton("📂")
        log_folder_btn.setFixedHeight(26)
        log_folder_btn.clicked.connect(self.open_log_folder)
        log_folder_btn.setToolTip("Открыть папку с логами")
        top_layout.addWidget(log_folder_btn)
        #change_folder_btn = QPushButton("⚙️")
        #change_folder_btn.setFixedHeight(26)
        #change_folder_btn.clicked.connect(self.change_save_folder)
        #change_folder_btn.setToolTip("Изменить папку для отчетов")
        #top_layout.addWidget(change_folder_btn)
        # Кнопка настроек
        settings_btn = QPushButton("⚙")
        settings_btn.setFixedHeight(26)
        settings_btn.clicked.connect(self.open_settings)
        settings_btn.setToolTip("Настройки")
        top_layout.addWidget(settings_btn)
        clear_logs_btn = QPushButton("🗑️")
        clear_logs_btn.setFixedHeight(26)
        clear_logs_btn.clicked.connect(self.clear_all_logs)
        clear_logs_btn.setToolTip("Очистить все логи и задачи")
        top_layout.addWidget(clear_logs_btn)
        main_layout.addLayout(top_layout)
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.tasks_container = QWidget()
        self.tasks_layout = QVBoxLayout(self.tasks_container)
        self.tasks_layout.addStretch()
        self.tasks_layout.setSpacing(5)
        scroll_area.setWidget(self.tasks_container)
        main_layout.addWidget(scroll_area)
        input_form = QFrame()
        input_form.setFrameShape(QFrame.Shape.StyledPanel)
        form_layout = QVBoxLayout(input_form)
        self.task_input = QLineEdit()
        self.task_input.setPlaceholderText("Задача")
        self.task_input.returnPressed.connect(self.handle_enter_pressed)
        self.link_input = QLineEdit()
        self.link_input.setPlaceholderText("Ссылка")
        self.link_input.returnPressed.connect(self.handle_enter_pressed)
        self.time_input = QLineEdit()
        self.time_input.setPlaceholderText("Время (минуты или 13:24-16:39)")
        self.time_input.returnPressed.connect(self.handle_enter_pressed)
        form_layout.addWidget(self.task_input)
        form_layout.addWidget(self.link_input)
        form_layout.addWidget(self.time_input)
        btn_layout = QHBoxLayout()
        add_btn = QPushButton("✅ Добавить задачу")
        add_btn.clicked.connect(self.add_task)
        stats_btn = QPushButton("📊 Статистика")
        stats_btn.clicked.connect(self.show_stats)
        save_btn = QPushButton("💾 Сохранить в Excel")
        save_btn.clicked.connect(self.save_excel_gui)
        quit_btn = QPushButton("🚪 Завершить день")
        quit_btn.clicked.connect(self.quit_app)
        btn_layout.addWidget(add_btn)
        btn_layout.addWidget(stats_btn)
        btn_layout.addWidget(save_btn)
        btn_layout.addWidget(quit_btn)
        form_layout.addLayout(btn_layout)
        main_layout.addWidget(input_form)

    def handle_enter_pressed(self):
        """Обработчик нажатия Enter в полях ввода."""
        sender = self.sender()
        if sender == self.task_input:
            self.add_task()
        elif sender == self.link_input:
            self.add_task()
        elif sender == self.time_input:
            self.add_task()

    def apply_theme(self):
        """Применяет темную тему к приложению."""
        old_move_event = self.moveEvent
        old_resize_event = self.resizeEvent
        self.moveEvent = lambda event: super(MainWindow, self).moveEvent(event)
        self.resizeEvent = lambda event: super(MainWindow, self).resizeEvent(event)
        try:
            palette = self.palette()
            palette.setColor(QPalette.ColorRole.Window, QColor(30, 30, 30))
            palette.setColor(QPalette.ColorRole.WindowText, QColor(220, 220, 220))
            palette.setColor(QPalette.ColorRole.Base, QColor(40, 40, 40))
            palette.setColor(QPalette.ColorRole.AlternateBase, QColor(50, 50, 50))
            palette.setColor(QPalette.ColorRole.ToolTipBase, QColor(25, 25, 25))
            palette.setColor(QPalette.ColorRole.ToolTipText, QColor(220, 220, 220))
            palette.setColor(QPalette.ColorRole.Text, QColor(220, 220, 220))
            palette.setColor(QPalette.ColorRole.Button, QColor(50, 50, 50))
            palette.setColor(QPalette.ColorRole.ButtonText, QColor(220, 220, 220))
            palette.setColor(QPalette.ColorRole.BrightText, QColor(255, 0, 0))
            palette.setColor(QPalette.ColorRole.Link, QColor(42, 130, 218))
            palette.setColor(QPalette.ColorRole.Highlight, QColor(42, 130, 218))
            palette.setColor(QPalette.ColorRole.HighlightedText, QColor(30, 30, 30))
            self.setStyleSheet(f"""
                QMainWindow, QWidget {{ background-color: #1e1e1e; color: #dcdcdc; }}
                QLineEdit, QTextEdit {{ background-color: #2a2a2a; color: #dcdcdc; border: 1px solid #555; padding: 5px; }}
                QPushButton {{ 
                    background-color: #3a3a3a; 
                    color: #dcdcdc; 
                    border: 1px solid #555; 
                    padding: 4px 8px;
                    margin: 2px;
                    border-radius: 4px;
                }}
                QPushButton:hover {{ background-color: #4a4a4a; }}
                QScrollArea {{ border: none; }}
                QLabel {{ color: #dcdcdc; }}
                QComboBox {{ background-color: #2a2a2a; color: #dcdcdc; border: 1px solid #555; }}
                /* === Стиль скроллбара === */
                QScrollBar:vertical {{
                    background: transparent;
                    width: 8px;
                    margin: 0px 0px 0px 0px;
                    border: none;
                }}
                QScrollBar::handle:vertical {{
                    background: rgba(100, 100, 100, 150);
                    border-radius: 4px;
                    min-height: 20px;
                }}
                QScrollBar::handle:vertical:hover {{
                    background: rgba(120, 120, 120, 200);
                }}
                /* Скрываем кнопки прокрутки */
                QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                    height: 0px;
                    width: 0px;
                    background: none;
                }}
                QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {{
                    background: none;
                }}
            """)
        finally:
            self.moveEvent = old_move_event
            self.resizeEvent = old_resize_event

    def reload_all_task_widgets(self):
        """Перезагружает все виджеты задач."""
        for i in reversed(range(self.tasks_layout.count() - 1)):
            widget = self.tasks_layout.itemAt(i).widget()
            if widget:
                widget.deleteLater()
        self.load_tasks_to_ui()

    def load_tasks_to_ui(self):
        """Загружает задачи в пользовательский интерфейс."""
        # Очищаем текущие виджеты
        for i in reversed(range(self.tasks_layout.count() - 1)):
            widget = self.tasks_layout.itemAt(i).widget()
            if widget:
                widget.deleteLater()
        self.tasks_widgets.clear()
        # Разделяем задачи на завершенные и активные
        completed_tasks = [(idx, task) for idx, task in enumerate(tasks) if "timer_start" not in task]
        active_tasks = [(idx, task) for idx, task in enumerate(tasks) if "timer_start" in task]
        # Сначала отображаем завершенные задачи
        for idx, task in completed_tasks:
            task_widget = TaskItemWidget(task, idx, self)
            insert_pos = max(0, self.tasks_layout.count() - 1)
            self.tasks_layout.insertWidget(insert_pos, task_widget)
            self.tasks_widgets.append(task_widget)
        # Затем отображаем активные задачи
        for idx, task in active_tasks:
            task_widget = TaskItemWidget(task, idx, self)
            insert_pos = max(0, self.tasks_layout.count() - 1)
            self.tasks_layout.insertWidget(insert_pos, task_widget)
            self.tasks_widgets.append(task_widget)
        # Прокручиваем вниз
        scroll_area = self.findChild(QScrollArea)
        if scroll_area:
            scroll_bar = scroll_area.verticalScrollBar()
            QTimer.singleShot(50, lambda: scroll_bar.setValue(scroll_bar.maximum()))

    def add_task(self):
        """Добавляет новую задачу."""
        task_text = self.task_input.text().strip()
        if not task_text:
            QMessageBox.warning(self, "Ошибка", "Введите задачу!")
            return
        is_completed = task_text.startswith("!")
        if is_completed:
            task_text = task_text[1:].strip()
        task_text, extracted_link = extract_markdown_links(task_text)
        link = self.link_input.text().strip() or extracted_link
        time_str = self.time_input.text().strip()
        start_timestamp = time.time()
        end_timestamp = None
        if time_str:
            if "-" in time_str:
                # Попробуем обработать оба формата: HH:MM-HH:MM и HHMM-HHMM
                try:
                    # Проверяем формат HHMM-HHMM (без двоеточий)
                    if len(time_str.replace(':', '')) == 9 and time_str.count('-') == 1:
                        start_part, end_part = time_str.split('-')
                        if len(start_part) == 4 and len(end_part) == 4:
                            h1 = int(start_part[:2])
                            m1 = int(start_part[2:])
                            h2 = int(end_part[:2])
                            m2 = int(end_part[2:])
                            # Проверяем корректность времени
                            if 0 <= h1 <= 23 and 0 <= m1 <= 59 and 0 <= h2 <= 23 and 0 <= m2 <= 59:
                                today = datetime.date.today()
                                start_dt = datetime.datetime(today.year, today.month, today.day, h1, m1)
                                end_dt = datetime.datetime(today.year, today.month, today.day, h2, m2)
                                if end_dt < start_dt:
                                    end_dt += datetime.timedelta(days=1)
                                # Вычисляем разницу в минутах
                                diff_seconds = (end_dt - start_dt).total_seconds()
                                minutes = diff_seconds / 60
                                # Проверяем адекватность продолжительности
                                if minutes > 24 * 60:
                                    raise ValueError("Продолжительность задачи не может превышать 24 часа")
                                start_timestamp = start_dt.timestamp()
                                end_timestamp = end_dt.timestamp()
                                # Форматируем для отображения в стандартный вид
                                time_str = f"{h1:02d}:{m1:02d}-{h2:02d}:{m2:02d}"
                            else:
                                raise ValueError("Некорректное время в формате HHMM")
                        else:
                            raise ValueError("Некорректная длина частей в формате HHMM-HHMM")
                    else:
                        # Стандартный формат HH:MM-HH:MM
                        start_str, end_str = time_str.split("-")
                        if len(start_str.split(":")) != 2 or len(end_str.split(":")) != 2:
                            raise ValueError("Неверный формат диапазона времени")
                        h1, m1 = map(int, start_str.split(":"))
                        h2, m2 = map(int, end_str.split(":"))
                        # Проверяем корректность времени
                        if not (0 <= h1 <= 23 and 0 <= m1 <= 59 and 0 <= h2 <= 23 and 0 <= m2 <= 59):
                            raise ValueError("Некорректное время")
                        today = datetime.date.today()
                        start_dt = datetime.datetime(today.year, today.month, today.day, h1, m1)
                        end_dt = datetime.datetime(today.year, today.month, today.day, h2, m2)
                        if end_dt < start_dt:
                            end_dt += datetime.timedelta(days=1)
                        # Вычисляем разницу в минутах
                        diff_seconds = (end_dt - start_dt).total_seconds()
                        minutes = diff_seconds / 60
                        # Проверяем адекватность продолжительности
                        if minutes > 24 * 60:
                            raise ValueError("Продолжительность задачи не может превышать 24 часа")
                        start_timestamp = start_dt.timestamp()
                        end_timestamp = end_dt.timestamp()
                except Exception as e:
                    QMessageBox.warning(self, "Ошибка", f"Неверный формат диапазона времени или время вне допустимых пределов: {str(e)}\nИспользуйте формат HH:MM-HH:MM (например, 13:24-16:39) или HHMM-HHMM (например, 1324-1639)")
                    return
            else:
                # Обработка времени в минутах
                try:
                    minutes = float(time_str)
                    # Проверяем адекватность продолжительности
                    if minutes <= 0:
                        raise ValueError("Время должно быть положительным")
                    if minutes > 24 * 60:  # Более 24 часов
                        raise ValueError("Время не может превышать 24 часа (1440 минут)")
                except ValueError as e:
                    QMessageBox.warning(self, "Ошибка", f"Неверный формат времени: {str(e)}\nВведите положительное число минут (не более 1440) или диапазон времени в формате HH:MM-HH:MM или HHMM-HHMM")
                    return
                # Для времени в минутах: начало - текущее время, конец - начало + минуты
                start_timestamp = time.time()
                end_timestamp = start_timestamp + (minutes * 60)
            hours_hundredths = round(minutes / 60, 2)
            display_time_str = "<1 минуты" if minutes < 1 else f"{int(minutes)}:{int((minutes%1)*60):02d}"
            if "-" in time_str:
                display_time_str = time_str
            tasks.append({
                "id": str(uuid.uuid4()),
                "task": task_text,
                "link": link,
                "time_str": display_time_str,
                "minutes": round(minutes, 2),
                "hours_hundredths": hours_hundredths,
                "start_timestamp": start_timestamp,
                "end_timestamp": end_timestamp
            })
        else:
            # Активная задача
            tasks.append({
                "id": str(uuid.uuid4()),
                "task": task_text,
                "link": link,
                "timer_start": time.time(),
                "start_timestamp": start_timestamp,
                "paused_total": 0.0, # Общее время в паузе
                "pause_history": [], # История пауз
                "minutes": 0.0,
                "hours_hundredths": 0.0,
                "time_str": "0:00"
            })
        save_backup()
        new_index = len(tasks) - 1
        task_widget = TaskItemWidget(tasks[new_index], new_index, self)
        insert_pos = max(0, self.tasks_layout.count() - 1)
        self.tasks_layout.insertWidget(insert_pos, task_widget)
        self.tasks_widgets.append(task_widget)
        scroll_area = self.findChild(QScrollArea)
        if scroll_area:
            scroll_bar = scroll_area.verticalScrollBar()
            QTimer.singleShot(50, lambda: scroll_bar.setValue(scroll_bar.maximum()))
        self.task_input.clear()
        self.link_input.clear()
        self.time_input.clear()

    def edit_task(self, index):
        """Открывает диалог редактирования задачи."""
        if index < 0 or index >= len(tasks):
            return
        task = tasks[index]
        dialog = EditTaskDialog(task, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            new_data = dialog.get_data()
            if new_data:
                # Сохраняем старый ID
                old_id = tasks[index].get("id", str(uuid.uuid4()))
                # Обновляем данные задачи
                tasks[index].update(new_data)
                # Восстанавливаем ID
                tasks[index]["id"] = old_id
                # Если это активная задача, обновляем timer_start
                if "timer_start" in tasks[index]:
                    tasks[index]["timer_start"] = tasks[index]["start_timestamp"]
                save_backup()
                self.load_tasks_to_ui()

    def delete_task(self, index):
        """Удаляет задачу."""
        if index < 0 or index >= len(tasks):
            return
        reply = QMessageBox.question(self, "Подтверждение", "Вы уверены, что хотите удалить эту задачу?",
                                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            tasks.pop(index)
            save_backup()
            self.load_tasks_to_ui()

    def show_stats(self):
        """Показывает статистику по задачам."""
        total_tasks = len(tasks)
        total_minutes = round(sum(t.get('minutes', 0) for t in tasks), 2)
        total_hours = round(sum(t.get('hours_hundredths', 0) for t in tasks), 2)

        # Подсчет активных и приостановленных задач
        active_tasks_count = 0
        paused_tasks_count = 0
        active_tasks_time = 0.0
        paused_tasks_time = 0.0

        for t in tasks:
            if "timer_start" in t:
                if t.get("is_paused", False):
                    paused_tasks_count += 1
                    # Для задач на паузе считаем накопленное время
                    paused_total = t.get("paused_total", 0.0)
                    elapsed = time.time() - t["timer_start"] - paused_total
                    paused_tasks_time += elapsed / 60
                else:
                    active_tasks_count += 1
                    # Для активных задач считаем текущее время
                    paused_total = t.get("paused_total", 0.0)
                    elapsed = time.time() - t["timer_start"] - paused_total
                    active_tasks_time += elapsed / 60

        link_minutes = {}
        for t in tasks:
            if t.get('link'):
                d = parse_domain(t['link'])
                link_minutes[d] = link_minutes.get(d, 0) + t.get('minutes', 0)
        top_links = sorted(link_minutes.items(), key=lambda x: x[1], reverse=True)[:3]

        task_minutes = {}
        for t in tasks:
            task_minutes[t['task']] = task_minutes.get(t['task'], 0) + t.get('minutes', 0)
        top_tasks = sorted(task_minutes.items(), key=lambda x: x[1], reverse=True)[:3]

        stats_parts = []
        stats_parts.append("<b>📊 ОБЩАЯ СТАТИСТИКА</b>")
        stats_parts.append(f"  Всего задач: <b>{total_tasks}</b>")
        stats_parts.append(f"  Общее время: <b>{total_minutes} мин</b> / <b>{total_hours} ч</b>")

        # Добавляем информацию о текущих задачах
        if active_tasks_count > 0 or paused_tasks_count > 0:
            stats_parts.append("")
            stats_parts.append("<b>Текущие задачи:</b>")
            if active_tasks_count > 0:
                stats_parts.append(f"  В работе: <b>{active_tasks_count}</b> задач ({round(active_tasks_time, 2)} мин)")
            if paused_tasks_count > 0:
                stats_parts.append(f"  На паузе: <b>{paused_tasks_count}</b> задач ({round(paused_tasks_time, 2)} мин)")

        if top_links:
            stats_parts.append("")
            stats_parts.append("<b>ТОП-3 ССЫЛОК ПО ВРЕМЕНИ</b>")
            for domain, mins in top_links:
                stats_parts.append(f"  🔗 {domain}: {round(mins, 2)} мин")
        if top_tasks:
            stats_parts.append("")
            stats_parts.append("<b>ТОП-3 ЗАДАЧ ПО ВРЕМЕНИ</b>")
            for name, mins in top_tasks:
                stats_parts.append(f"  📝 {name}: {round(mins, 2)} мин")

        stats_text = "<br>".join(stats_parts)
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("Статистика")
        msg_box.setTextFormat(Qt.TextFormat.RichText)
        msg_box.setText(stats_text)
        msg_box.exec()

    def save_excel_gui(self):
        """Сохраняет данные в Excel-файл через графический интерфейс."""
        global SAVE_DIR
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Сохранить отчет",
            os.path.join(SAVE_DIR, f"{datetime.date.today().strftime('%Y-%m-%d')}.xlsx"),
            "Excel Files (*.xlsx);;All Files (*)"
        )
        if file_path:
            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Tasks"
                # Добавляем новые заголовки
                ws.append(["Задача", "Ссылка (домен)", "Время (мин:сек)", "Время (часы в сотых)", "Занято, мин", "Период выполнения"])
                total_minutes = 0.0
                total_hours_hundredths = 0.0
                total_occupied_minutes = 0.0 # Для подсчета итога по "Занято, мин"
                keys = [key_for_group(t) for t in tasks]
                dup_keys = {k for k, c in Counter(keys).items() if c > 1}
                row_to_key = {}
                for t in tasks:
                    domain = parse_domain(t['link']) if t['link'] else ""
                    row_index = ws.max_row + 1
                    ws.cell(row=row_index, column=1, value=t['task'])
                    if t['link']:
                        link_cell = ws.cell(row=row_index, column=2, value=domain)
                        link_cell.hyperlink = t['link']
                        link_cell.style = "Hyperlink"
                    else:
                        ws.cell(row=row_index, column=2, value="")
                    ws.cell(row=row_index, column=3, value=t['time_str'])
                    ws.cell(row=row_index, column=4, value=t['hours_hundredths'])
                    # Добавляем новые колонки
                    # "Занято, мин" - это просто minutes, округленные до 2 знаков
                    occupied_minutes = round(t.get('minutes', 0), 2)
                    ws.cell(row=row_index, column=5, value=occupied_minutes)
                    # "Период выполнения"
                    period_str = format_time_period(t.get('start_timestamp'), t.get('end_timestamp'))
                    ws.cell(row=row_index, column=6, value=period_str)
                    k = key_for_group(t)
                    if k in dup_keys:
                        row_to_key[row_index] = k
                    total_minutes += t['minutes']
                    total_hours_hundredths += t['hours_hundredths']
                    total_occupied_minutes += occupied_minutes # Суммируем для итога
                n_tasks_rows = len(tasks) + 1
                # Обновляем строку итогов
                ws.append([])
                ws.append(["ИТОГО", "", f"{round(total_minutes, 2)} мин", f"{round(total_hours_hundredths, 2)} ч", f"{round(total_occupied_minutes, 2)} мин", ""])
                grouped = group_tasks()
                if grouped:
                    ws.append([])
                    ws.append(["СВОДКА (объединено по совпадению ссылки или названия)"])
                    for gt in grouped:
                        ws.append([
                            gt["task"],
                            parse_domain(gt["link"]) if gt["link"] else "",
                            f"{round(gt['minutes'], 2)} мин",
                            f"{round(gt['hours_hundredths'], 2)} ч",
                            "", # Пусто для сводки
                            ""  # Пусто для сводки
                        ])
                # Передаем n_tasks_rows+1, так как добавили 2 колонки
                style_tasks_and_summary(
                    ws=ws,
                    n_tasks_rows=n_tasks_rows,
                    grouped=grouped,
                    group_key_to_rows={"row_to_key": row_to_key}
                )
                wb.save(file_path)
                new_save_dir = os.path.dirname(file_path)
                if new_save_dir != SAVE_DIR:
                    SAVE_DIR = new_save_dir
                    app_config['USER']['save_dir'] = SAVE_DIR
                    save_settings(app_config)
                QMessageBox.information(self, "Успех", f"Файл успешно сохранён: {file_path}")
                reply = QMessageBox.question(self, "Открыть папку?", "Открыть папку с файлом?",
                                            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
                if reply == QMessageBox.StandardButton.Yes:
                    if sys.platform == "win32":
                        os.startfile(os.path.dirname(file_path))
                    elif sys.platform == "darwin":
                        os.system(f'open "{os.path.dirname(file_path)}"')
                    else:
                        os.system(f'xdg-open "{os.path.dirname(file_path)}"')
            except PermissionError:
                QMessageBox.critical(self, "Ошибка", "Файл занят другим процессом. Пожалуйста, закройте файл и попробуйте снова.")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить файл: {str(e)}")

    def open_log_folder(self):
        """Открывает папку с логами."""
        if os.path.exists(LOG_DIR):
            if sys.platform == "win32":
                os.startfile(LOG_DIR)
            elif sys.platform == "darwin":
                os.system(f'open "{LOG_DIR}"')
            else:
                os.system(f'xdg-open "{LOG_DIR}"')
        else:
            QMessageBox.warning(self, "Папка не найдена", "Временная папка с логами ещё не создана.")

    def change_save_folder(self):
        """Изменяет папку для сохранения отчетов."""
        global SAVE_DIR, app_config
        folder = QFileDialog.getExistingDirectory(
            self,
            "Выберите папку для сохранения отчётов Excel",
            SAVE_DIR
        )
        if folder:
            SAVE_DIR = folder
            app_config['USER']['save_dir'] = SAVE_DIR
            save_settings(app_config)
            QMessageBox.information(self, "Успех", f"Папка сохранения изменена на:\n{SAVE_DIR}")

    def open_settings(self):
        """Открывает диалог настроек."""
        global SAVE_DIR, LOG_DIR  # Добавляем глобальные переменные
        dialog = SettingsDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            settings_data = dialog.get_data()
            # Обновляем настройки
            SAVE_DIR = settings_data["save_dir"]
            LOG_DIR = settings_data["log_dir"]
            # Сохраняем настройки
            app_config['USER']['save_dir'] = SAVE_DIR
            # app_config['USER']['log_dir'] = LOG_DIR # Лог-директория обычно не изменяется
            save_settings(app_config)
            QMessageBox.information(self, "Успех", "Настройки сохранены.")

    def clear_all_logs(self):
        """Очищает все логи и задачи."""
        reply = QMessageBox.question(
            self,
            "Подтверждение",
            "Вы уверены, что хотите удалить ВСЕ логи и сбросить список задач?\n"
            "Это действие необратимо.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        try:
            if os.path.exists(LOG_DIR):
                for file in os.listdir(LOG_DIR):
                    if file.endswith(".txt"):
                        os.remove(os.path.join(LOG_DIR, file))
            if os.path.exists(SAVE_DIR):
                for file in os.listdir(SAVE_DIR):
                    if file.endswith(".xlsx") and file.startswith(datetime.date.today().strftime("%Y-%m-%d")):
                        os.remove(os.path.join(SAVE_DIR, file))
            tasks.clear()
            save_backup()
            self.load_tasks_to_ui()
            QMessageBox.information(self, "Очистка", "Все логи и задачи удалены.")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось очистить логи:\n{str(e)}")

    def quit_app(self):
        """Завершает работу приложения."""
        reply = QMessageBox.question(
            self,
            "Завершение дня",
            "Хотите сохранить отчет перед выходом?\nВсе активные задачи будут завершены и сохранены.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No | QMessageBox.StandardButton.Cancel
        )
        if reply == QMessageBox.StandardButton.Yes:
            # Завершаем все активные задачи перед сохранением
            self.finish_all_active_tasks()
            self.save_excel_gui()
            save_backup()
            if os.path.exists(active_tasks_file):
                try:
                    os.remove(active_tasks_file)
                except Exception as e:
                    print(Fore.YELLOW + f"Не удалось удалить временный файл: {e}")
            QApplication.quit()
        elif reply == QMessageBox.StandardButton.No:
            # Завершаем все активные задачи перед выходом
            self.finish_all_active_tasks()
            save_backup()
            if os.path.exists(active_tasks_file):
                try:
                    os.remove(active_tasks_file)
                except Exception as e:
                    print(Fore.YELLOW + f"Не удалось удалить временный файл: {e}")
            QApplication.quit()
        else:
            pass

    def finish_all_active_tasks(self):
        """Завершает все активные задачи и сохраняет их как выполненные."""
        # Создаем список индексов активных задач (в обратном порядке для корректного удаления)
        active_indices = [i for i, task in enumerate(tasks) if "timer_start" in task]
        active_indices.reverse()
        # Завершаем каждую активную задачу
        for index in active_indices:
            finished_task = finish_active_task(index)
            if finished_task:
                # Удаляем активную задачу из временного файла
                remove_active_task(index)
                # Заменяем активную задачу на завершенную
                tasks[index] = finished_task
        # Перезагружаем интерфейс
        self.load_tasks_to_ui()

    def on_window_destroyed(self):
        """Вызывается при закрытии окна программы."""
        # Завершаем все активные задачи перед закрытием
        self.finish_all_active_tasks()
        save_settings(app_config)
        save_backup()
        # Удаляем временный файл активных задач, так как все задачи уже завершены
        if os.path.exists(active_tasks_file):
            try:
                os.remove(active_tasks_file)
            except Exception as e:
                print(Fore.YELLOW + f"Не удалось удалить временный файл при закрытии: {e}")

    def moveEvent(self, event):
        """Обработчик перемещения окна."""
        self.save_window_state()
        super().moveEvent(event)

    def resizeEvent(self, event):
        """Обработчик изменения размера окна."""
        self.save_window_state()
        super().resizeEvent(event)

    def save_window_state(self):
        """Сохраняет состояние окна (позицию и размер)."""
        try:
            app_config['USER']['window_x'] = str(self.x())
            # Костыль: сохраняем window_y со смещением -31 пиксель
            app_config['USER']['window_y'] = str(self.y() + 61)
            app_config['USER']['window_width'] = str(self.width())
            app_config['USER']['window_height'] = str(self.height())
            save_settings(app_config)
        except Exception as e:
            import traceback
            with open("error_log.txt", "a", encoding='utf-8') as logf:
                logf.write(f"[ERROR_save_window_state] {e}\n{traceback.format_exc()}\n")

# Точка входа в приложение
if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())