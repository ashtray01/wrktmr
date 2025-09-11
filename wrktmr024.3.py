import os
import sys
import time
import datetime
import signal
import re
import tempfile
import ctypes
import uuid
import configparser
from urllib.parse import urlparse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter
from colorama import init, Fore, Style

def set_dark_title_bar_qt(window):
    if sys.platform != "win32":
        return
    try:
        hwnd = int(window.winId())
        dark_mode_attribute = 20
        value = ctypes.c_int(1)
        ctypes.windll.dwmapi.DwmSetWindowAttribute(
            hwnd,
            dark_mode_attribute,
            ctypes.byref(value),
            ctypes.sizeof(value)
        )
    except Exception as e:
        print(f"[Тёмный заголовок] Не удалось применить: {e}")

from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                             QTextEdit, QLineEdit, QPushButton, QLabel, QScrollArea,
                             QFrame, QMessageBox, QFileDialog, QComboBox, QDialog,
                             QDialogButtonBox, QFormLayout, QSpinBox, QDoubleSpinBox)
from PyQt6.QtCore import Qt, QTimer, QUrl, QTime, QSize, QPoint
from PyQt6.QtGui import QFont, QDesktopServices, QColor, QTextCharFormat, QTextCursor, QPalette, QIcon

init(autoreset=True)

def resource_path(relative_path):
    """ Получает абсолютный путь к ресурсу, работает для dev и для PyInstaller """
    try:
        # PyInstaller создаёт временную папку и сохраняет путь в _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

def get_config_dir():
    temp_dir = tempfile.gettempdir()
    return os.path.join(temp_dir, "wrktmr-tmp")

def get_log_dir():
    config_dir = get_config_dir()
    today = datetime.date.today()
    return os.path.join(config_dir, "logs", str(today.year), f"{today.month:02d}", f"{today.day:02d}")

CONFIG_FILE = os.path.join(get_config_dir(), "settings.ini")
_save_settings_lock = False

def load_settings():
    config = configparser.ConfigParser(interpolation=None)
    if not os.path.exists(CONFIG_FILE):
        os.makedirs(os.path.dirname(CONFIG_FILE), exist_ok=True)
        config['USER'] = {
            'save_dir': os.path.join(os.path.expanduser("~"), "Desktop"),
            'window_x': '100',
            'window_y': '100',
            'window_width': '900',
            'window_height': '700'
        }
        try:
            with open(CONFIG_FILE, 'w', encoding='utf-8') as configfile:
                config.write(configfile)
        except Exception as e:
            print(f"[ERROR] Не удалось создать файл настроек '{CONFIG_FILE}': {e}")
    else:
        try:
            config.read(CONFIG_FILE, encoding='utf-8')
            if 'USER' not in config:
                config['USER'] = {
                    'save_dir': os.path.join(os.path.expanduser("~"), "Desktop"),
                    'window_x': '100',
                    'window_y': '100',
                    'window_width': '900',
                    'window_height': '700'
                }
                save_settings(config)
        except Exception as e:
            print(f"[ERROR] Не удалось прочитать файл настроек '{CONFIG_FILE}': {e}")
            config['USER'] = {
                'save_dir': os.path.join(os.path.expanduser("~"), "Desktop"),
                'window_x': '100',
                'window_y': '100',
                'window_width': '900',
                'window_height': '700'
            }
    return config

def save_settings(config):
    try:
        with open(CONFIG_FILE, 'w', encoding='utf-8') as configfile:
            config.write(configfile)
        #print(f"[DEBUG] Настройки успешно сохранены в '{CONFIG_FILE}'.")
    except Exception as e:
        print(f"[ERROR] Не удалось сохранить настройки в '{CONFIG_FILE}': {e}")

app_config = load_settings()

if 'USER' not in app_config:
    app_config['USER'] = {}

user_section = app_config['USER']
SAVE_DIR = user_section.get('save_dir', os.path.join(os.path.expanduser("~"), "Desktop"))

default_user_settings = {
    'save_dir': os.path.join(os.path.expanduser("~"), "Desktop"),
    'window_x': '100',
    'window_y': '100',
    'window_width': '900',
    'window_height': '700'
}
for key, default_value in default_user_settings.items():
    if key not in user_section:
        user_section[key] = default_value

save_settings(app_config)

LOG_DIR = get_log_dir()
os.makedirs(LOG_DIR, exist_ok=True)
tasks = []
active_tasks_file = os.path.join(LOG_DIR, "active_tasks.tmp")
active_save_timer = None

def signal_handler(sig, frame):
    print(Fore.YELLOW + "\nОбнаружено Ctrl+C! Сохраняем данные...")
    save_backup()
    print(Fore.MAGENTA + "Данные сохранены. Программа завершена.")
    sys.exit(0)

signal.signal(signal.SIGINT, signal_handler)

def parse_domain(url):
    try:
        parsed = urlparse(url)
        return parsed.netloc or url
    except:
        return url

def format_time_range(start_timestamp, end_timestamp=None):
    try:
        if not isinstance(start_timestamp, (int, float)):
            start_str = "Начало: неизвестно"
        else:
            start_dt = datetime.datetime.fromtimestamp(start_timestamp)
            start_str = start_dt.strftime("%H:%M")
        if end_timestamp is None:
            return f"Начало {start_str} - В процессе"
        else:
            end_dt = datetime.datetime.fromtimestamp(end_timestamp)
            end_str = end_dt.strftime("%H:%M")
            return f"Начало {start_str} - Завершена {end_str}"
    except Exception as e:
        print(f"Ошибка форматирования времени: {e}")
        return "Время неизвестно"

def key_for_group(t):
    if t.get("link"):
        return t["link"].strip().lower()
    return t["task"].strip().lower()

def load_backup():
    today = datetime.date.today().strftime("%Y-%m-%d")
    log_file = os.path.join(LOG_DIR, f"{today}.txt")
    if os.path.exists(log_file):
        with open(log_file, "r", encoding="utf-8") as f:
            for line in f:
                try:
                    parts = line.strip().split(" | ")
                    if len(parts) < 4:
                        continue
                    task, link, time_str, hours_hundredths = parts
                    if time_str == "<1 минуты":
                        minutes = 0.5
                    else:
                        if "-" in time_str:
                            minutes = parse_time_range(time_str)
                        else:
                            try:
                                mins, secs = map(int, time_str.split(":"))
                                minutes = mins + secs / 60
                            except:
                                minutes = float(time_str) if time_str.replace('.','',1).isdigit() else 0.0
                    tasks.append({
                        "id": str(uuid.uuid4()),
                        "task": task,
                        "link": link,
                        "time_str": time_str,
                        "minutes": round(minutes, 2),
                        "hours_hundredths": float(hours_hundredths)
                    })
                except Exception as e:
                    print(Fore.RED + f"Ошибка чтения backup: {e}")
        if tasks:
            print(Fore.GREEN + f"Загружено {len(tasks)} задач из backup: {log_file}")
    active_ids = {t.get("id") for t in tasks if "timer_start" in t}
    if active_ids:
        tasks[:] = [t for t in tasks if "timer_start" in t or t.get("id") not in active_ids]

def load_active_tasks():
    if not os.path.exists(active_tasks_file):
        return
    try:
        with open(active_tasks_file, "r", encoding="utf-8") as f:
            for line in f:
                try:
                    parts = line.strip().split(" | ")
                    if len(parts) < 4:
                        continue
                    task_id, task, link, start_timestamp_str = parts
                    start_timestamp = float(start_timestamp_str)
                    tasks.append({
                        "id": task_id,
                        "task": task,
                        "link": link,
                        "timer_start": start_timestamp,
                        "start_timestamp": start_timestamp,
                        "minutes": 0.0,
                        "hours_hundredths": 0.0,
                        "time_str": "0:00"
                    })
                except Exception as e:
                    print(Fore.RED + f"Ошибка загрузки активной задачи: {e}")
        if tasks:
            active_count = len([t for t in tasks if 'timer_start' in t])
            print(Fore.GREEN + f"Загружено {active_count} активных задач из временного файла.")
    except Exception as e:
        print(Fore.RED + f"Ошибка чтения активных задач: {e}")

def save_backup():
    today = datetime.date.today().strftime("%Y-%m-%d")
    log_file = os.path.join(LOG_DIR, f"{today}.txt")
    os.makedirs(os.path.dirname(log_file), exist_ok=True)
    with open(log_file, "w", encoding="utf-8") as f:
        for t in tasks:
            if "timer_start" in t:
                continue
            clean_t = {k: v for k, v in t.items() if k != 'timer_start'}
            f.write(f"{clean_t['task']} | {clean_t['link']} | {clean_t['time_str']} | {clean_t['hours_hundredths']}\n")

def save_active_tasks():
    try:
        active_tasks = [t for t in tasks if "timer_start" in t]
        if not active_tasks:
            if os.path.exists(active_tasks_file):
                os.remove(active_tasks_file)
            return
        with open(active_tasks_file, "w", encoding="utf-8") as f:
            for t in active_tasks:
                task_copy = t.copy()
                start_time = task_copy.pop("timer_start")
                task_id = task_copy.get("id", str(uuid.uuid4()))
                f.write(f"{task_id} | {task_copy['task']} | {task_copy['link']} | {start_time}\n")
    except Exception as e:
        print(Fore.RED + f"Ошибка сохранения активных задач: {e}")

def remove_active_task(index):
    try:
        if index < 0 or index >= len(tasks):
            return
        task_id = tasks[index].get("id")
        if not task_id:
            return
        if not os.path.exists(active_tasks_file):
            return
        remaining_lines = []
        removed = False
        with open(active_tasks_file, "r", encoding="utf-8") as f:
            for line in f:
                parts = line.strip().split(" | ")
                if len(parts) < 4:
                    remaining_lines.append(line.strip())
                    continue
                saved_id, _, _, _ = parts
                if saved_id != task_id:
                    remaining_lines.append(line.strip())
                else:
                    removed = True
        with open(active_tasks_file, "w", encoding="utf-8") as f:
            for line in remaining_lines:
                f.write(line + "\n")
        if not remaining_lines:
            os.remove(active_tasks_file)
        if removed:
            print(Fore.GREEN + f"Активная задача {task_id} удалена из временного файла.")
    except Exception as e:
        print(Fore.RED + f"Ошибка удаления активной задачи: {e}")

def group_tasks():
    grouped = {}
    for t in tasks:
        link_key = t["link"].strip().lower() if t.get("link") else None
        if link_key:
            if link_key not in grouped:
                grouped[link_key] = {
                    "task": t["task"],
                    "link": t["link"],
                    "minutes": 0.0,
                    "hours_hundredths": 0.0,
                    "count": 0,
                    "key": link_key
                }
            grouped[link_key]["minutes"] += t["minutes"]
            grouped[link_key]["hours_hundredths"] += t["hours_hundredths"]
            grouped[link_key]["count"] += 1
        name_key = t["task"].strip().lower()
        if name_key not in grouped:
            grouped[name_key] = {
                "task": t["task"],
                "link": t["link"],
                "minutes": 0.0,
                "hours_hundredths": 0.0,
                "count": 0,
                "key": name_key
            }
        grouped[name_key]["minutes"] += t["minutes"]
        grouped[name_key]["hours_hundredths"] += t["hours_hundredths"]
        grouped[name_key]["count"] += 1
    return [g for g in grouped.values() if g["count"] > 1]

def parse_time_range(time_str):
    try:
        start_str, end_str = time_str.split("-")
        h1, m1 = map(int, start_str.split(":"))
        h2, m2 = map(int, end_str.split(":"))
        start_minutes = h1 * 60 + m1
        end_minutes = h2 * 60 + m2
        if end_minutes < start_minutes:
            end_minutes += 24 * 60
        diff = end_minutes - start_minutes
        return diff
    except:
        return 0.0

def extract_markdown_links(text):
    pattern = r'\[([^\]]+)\]\(([^)]+)\)'
    matches = re.findall(pattern, text)
    if matches:
        display_text = matches[0][0]
        url = matches[0][1]
        clean_text = re.sub(pattern, display_text, text, count=1)
        return clean_text.strip(), url.strip()
    return text.strip(), ""

def style_tasks_and_summary(ws, n_tasks_rows, grouped, group_key_to_rows):
    thin = Border(left=Side(style='thin'),
                  right=Side(style='thin'),
                  top=Side(style='thin'),
                  bottom=Side(style='thin'))
    palette = [
        "FFF2CC", "E2EFDA", "DDEBF7", "FCE4D6", "EDEDED",
        "CCE5FF", "FFD966", "C6E0B4", "D9E1F2", "FFEB9C",
        "E7E6E6", "D5E8D4", "D0CECE", "F8CBAD", "C9DAF8"
    ]
    def color_for_key(k):
        if not k:
            return None
        idx = abs(hash(k)) % len(palette)
        return PatternFill(start_color=palette[idx], end_color=palette[idx], fill_type="solid")
    max_row = ws.max_row
    max_col = ws.max_column
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for col in range(1, max_col + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(vertical="center", horizontal="left")
        c.fill = header_fill
        c.border = thin
    for r in range(2, 2 + n_tasks_rows - 1 if n_tasks_rows > 0 else 1):
        zebra = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid") if (r % 2 == 0) else None
        if r in group_key_to_rows["row_to_key"]:
            k = group_key_to_rows["row_to_key"][r]
            fill = color_for_key(k)
        else:
            fill = zebra
        for col in range(1, max_col + 1):
            cell = ws.cell(row=r, column=col)
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin
    total_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "ИТОГО":
            total_row = r
            break
    if total_row:
        for col in range(1, max_col + 1):
            c = ws.cell(row=total_row, column=col)
            c.font = Font(bold=True)
            c.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            c.border = thin
    summary_header_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "СВОДКА (объединено по совпадению ссылки или названия)":
            summary_header_row = r
            break
    if summary_header_row:
        for col in range(1, max_col + 1):
            c = ws.cell(row=summary_header_row, column=col)
            c.font = Font(bold=True)
            c.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            c.border = thin
        sr = summary_header_row + 1
        while sr <= ws.max_row:
            a = ws.cell(row=sr, column=1).value
            b = ws.cell(row=sr, column=2).value
            c3 = ws.cell(row=sr, column=3).value
            if a is None and b is None and c3 is None:
                break
            matched_key = None
            for g in grouped:
                if a and a == g["task"]:
                    matched_key = g["key"]
                    break
                if b and b == parse_domain(g["link"] or ""):
                    matched_key = g["key"]
                    break
            if matched_key:
                fill = color_for_key(matched_key)
                if fill:
                    for col in range(1, max_col + 1):
                        cell = ws.cell(row=sr, column=col)
                        cell.fill = fill
                        cell.border = thin
            sr += 1
    if n_tasks_rows > 0:
        ws.auto_filter.ref = f"A1:D{1 + n_tasks_rows}"
    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=max_col):
        for cell in row:
            if not cell.border:
                cell.border = thin
            if not cell.alignment:
                cell.alignment = Alignment(vertical="center")
    for col in range(1, max_col + 1):
        column = get_column_letter(col)
        max_len = 0
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[column].width = max(10, min(60, max_len + 2))

def save_excel():
    global SAVE_DIR
    today = datetime.date.today().strftime("%Y-%m-%d")
    file_name = os.path.join(SAVE_DIR, f"{today}.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tasks"
    ws.append(["Задача", "Ссылка (домен)", "Время (мин:сек)", "Время (часы в сотых)"])
    total_minutes = 0.0
    total_hours_hundredths = 0.0
    keys = [key_for_group(t) for t in tasks]
    dup_keys = {k for k, c in Counter(keys).items() if c > 1}
    row_to_key = {}
    for t in tasks:
        domain = parse_domain(t['link']) if t['link'] else ""
        row_index = ws.max_row + 1
        ws.cell(row=row_index, column=1, value=t['task'])
        if t['link']:
            link_cell = ws.cell(row=row_index, column=2, value=domain)
            link_cell.hyperlink = t['link']
            link_cell.style = "Hyperlink"
        else:
            ws.cell(row=row_index, column=2, value="")
        ws.cell(row=row_index, column=3, value=t['time_str'])
        ws.cell(row=row_index, column=4, value=t['hours_hundredths'])
        k = key_for_group(t)
        if k in dup_keys:
            row_to_key[row_index] = k
        total_minutes += t['minutes']
        total_hours_hundredths += t['hours_hundredths']
    n_tasks_rows = len(tasks) + 1
    ws.append([])
    ws.append(["ИТОГО", "", f"{round(total_minutes, 2)} мин", f"{round(total_hours_hundredths, 2)} ч"])
    grouped = group_tasks()
    if grouped:
        ws.append([])
        ws.append(["СВОДКА (объединено по совпадению ссылки или названия)"])
        for gt in grouped:
            ws.append([
                gt["task"],
                parse_domain(gt["link"]) if gt["link"] else "",
                f"{round(gt['minutes'], 2)} мин",
                f"{round(gt['hours_hundredths'], 2)} ч"
            ])
    style_tasks_and_summary(
        ws=ws,
        n_tasks_rows=n_tasks_rows,
        grouped=grouped,
        group_key_to_rows={"row_to_key": row_to_key}
    )
    try:
        wb.save(file_name)
        print(Fore.GREEN + f"Excel-отчёт сохранён: {file_name}")
        return file_name
    except Exception as e:
        print(Fore.RED + f"Ошибка при сохранении Excel: {e}")
        return None

class TaskItemWidget(QFrame):
    def __init__(self, task_data, index, parent=None):
        super().__init__(parent)
        self.task_data = task_data
        self.index = index
        self.parent_window = parent
        self.timer_label = None
        self.timer = None
        self.setup_ui()
        if "timer_start" in self.task_data:
            self.start_live_timer()

    def setup_ui(self):
        self.setStyleSheet("""
                    TaskItemWidget {
                        background-color: #2a2a2a;
                        border-radius: 6px;
                        border: 1px solid #444;
                        margin: 4px;
                    }
                """)
        self.setFrameShadow(QFrame.Shadow.Raised)
        main_layout = QHBoxLayout()
        main_layout.setContentsMargins(10, 8, 10, 8)
        main_layout.setSpacing(10)
        left_layout = QVBoxLayout()
        left_layout.setSpacing(3)
        task_label = QLabel()
        task_text = self.task_data['task']
        task_label.setTextFormat(Qt.TextFormat.RichText)
        task_label.setText(self.format_markdown_links(task_text))
        task_label.setWordWrap(True)
        task_label.setTextInteractionFlags(Qt.TextInteractionFlag.TextBrowserInteraction)
        task_label.linkActivated.connect(self.open_link)
        task_label.setStyleSheet("font-size: 13px;")
        left_layout.addWidget(task_label)
        if self.task_data['link']:
            domain = parse_domain(self.task_data['link'])
            link_label = QLabel(f"<a href='{self.task_data['link']}' style='color: #64b5f6; font-size: 11px;'>🔗 {domain}</a>")
            link_label.setTextFormat(Qt.TextFormat.RichText)
            link_label.setTextInteractionFlags(Qt.TextInteractionFlag.TextBrowserInteraction)
            link_label.linkActivated.connect(self.open_link)
            left_layout.addWidget(link_label)
        self.time_display = QLabel()
        self.update_time_display()
        self.time_display.setStyleSheet("font-size: 11px; color: #aaa;")
        left_layout.addWidget(self.time_display)
        main_layout.addLayout(left_layout)
        button_layout = QVBoxLayout()
        button_layout.setSpacing(5)
        button_layout.addStretch()
        button_width = int(self.parent_window.width() * 0.2) if self.parent_window else 150
        button_height = 32
        if "timer_start" in self.task_data:
            self.finish_btn = QPushButton("Завершить")
            self.finish_btn.setFixedSize(button_width, button_height)
            self.finish_btn.clicked.connect(self.confirm_finish_task)
            self.finish_btn.setStyleSheet("""
                QPushButton {
                    background-color: #d32f2f;
                    color: white;
                    border: 1px solid #b71c1c;
                    border-radius: 4px;
                    padding: 4px 8px;
                }
                QPushButton:hover {
                    background-color: #f44336;
                }
            """)
            button_layout.addWidget(self.finish_btn)
        else:
            edit_btn = QPushButton("✏️ Изменить")
            edit_btn.setFixedSize(button_width, button_height)
            edit_btn.clicked.connect(self.edit_task)
            delete_btn = QPushButton("🗑️ Удалить")
            delete_btn.setFixedSize(button_width, button_height)
            delete_btn.clicked.connect(self.delete_task)
            button_layout.addWidget(edit_btn)
            button_layout.addWidget(delete_btn)
        button_layout.addStretch()
        main_layout.addLayout(button_layout)
        self.setLayout(main_layout)
        self.setFixedHeight(90)

    def format_markdown_links(self, text):
        pattern = r'\[([^\]]+)\]\(([^)]+)\)'
        def repl(match):
            display = match.group(1)
            url = match.group(2)
            return f'<a href="{url}" style="color: #64b5f6;">{display}</a>'
        return re.sub(pattern, repl, text)

    def open_link(self, url):
        QDesktopServices.openUrl(QUrl(url))

    def update_time_display(self):
        if "timer_start" in self.task_data:
            elapsed = time.time() - self.task_data['timer_start']
            minutes = int(elapsed // 60)
            hours_hundredths = round(elapsed / 3600, 2)
            anim_chars = ['|', '/', '-', '\\']
            if not hasattr(self, '_anim_index'):
                self._anim_index = 0
            else:
                self._anim_index = (self._anim_index + 1) % 4
            anim_char = anim_chars[self._anim_index]
            time_range_str = format_time_range(self.task_data['start_timestamp'])
            self.time_display.setText(f"<span style='color: #4caf50;'>{minutes} мин ({hours_hundredths} ч) {time_range_str}</span> {anim_char}")
        else:
            minutes = int(self.task_data['minutes'])
            hours_hundredths = self.task_data['hours_hundredths']
            start_ts = self.task_data.get('start_timestamp')
            end_ts = self.task_data.get('end_timestamp')
            time_range_str = format_time_range(start_ts, end_ts) if start_ts else ""
            self.time_display.setText(f"{minutes} мин ({hours_hundredths} ч) {time_range_str}")

    def start_live_timer(self):
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.update_time_display)
        self.timer.start(1000)

    def confirm_finish_task(self):
        reply = QMessageBox.question(
            self,
            "Завершение задачи",
            "Вы уверены, что хотите завершить эту задачу?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.Yes:
            self.finish_task()

    def finish_task(self):
        if "timer_start" not in self.task_data:
            return
        remove_active_task(self.index)
        elapsed = time.time() - self.task_data['timer_start']
        minutes = elapsed / 60
        hours_hundredths = round(minutes / 60, 2)
        time_str = "<1 минуты" if minutes < 1 else f"{int(minutes)}:{int((minutes%1)*60):02d}"
        tasks[self.index].update({
            "minutes": round(minutes, 2),
            "hours_hundredths": hours_hundredths,
            "time_str": time_str,
            "end_timestamp": time.time()
        })
        del tasks[self.index]["timer_start"]
        if self.timer:
            self.timer.stop()
        save_backup()
        self.parent_window.load_tasks_to_ui()

    def edit_task(self):
        self.parent_window.edit_task(self.index)

    def delete_task(self):
        self.parent_window.delete_task(self.index)

class EditTaskDialog(QDialog):
    def __init__(self, task_data, parent=None):
        super().__init__(parent)
        self.task_data = task_data
        self.setWindowTitle("✏️ Редактировать задачу")
        self.resize(500, 200)
        self.setup_ui()

    def setup_ui(self):
        layout = QFormLayout()
        self.task_input = QLineEdit(self.task_data.get("task", ""))
        self.link_input = QLineEdit(self.task_data.get("link", ""))
        minutes_value = self.task_data.get("minutes", 0.0)
        self.time_input = QLineEdit(str(minutes_value))
        layout.addRow("📝 Задача:", self.task_input)
        layout.addRow("🔗 Ссылка:", self.link_input)
        layout.addRow("⏱️ Время (в минутах):", self.time_input)
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)
        self.setLayout(layout)

    def get_data(self):
        task_text = self.task_input.text().strip()
        link = self.link_input.text().strip()
        time_input_text = self.time_input.text().strip()
        minutes = 0.0
        try:
            minutes = float(time_input_text)
        except:
            QMessageBox.warning(self, "Ошибка", "Неверный формат времени. Введите число минут.")
            return None
        hours_hundredths = round(minutes / 60, 2)
        time_str = "<1 минуты" if minutes < 1 else f"{int(minutes)}:{int((minutes%1)*60):02d}"
        return {
            "task": task_text,
            "link": link,
            "time_str": time_str,
            "minutes": round(minutes, 2),
            "hours_hundredths": hours_hundredths
        }

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        x = int(app_config['USER'].get('window_x', 100))
        y = int(app_config['USER'].get('window_y', 100))
        width = int(app_config['USER'].get('window_width', 900))
        height = int(app_config['USER'].get('window_height', 700))
        self.setGeometry(x, y, width, height)
        self.setWindowTitle("wrktmr v024.3")
        icon_path = resource_path(os.path.join('icon', 'icon.ico'))
        if os.path.exists(icon_path):
            icon = QIcon(icon_path)
            self.setWindowIcon(icon)
            QApplication.instance().setWindowIcon(icon) 
        else:
            print(f"[DEBUG] Иконка не найдена по пути: {icon_path}")
        self.tasks_widgets = []
        load_active_tasks()
        load_backup()
        self.init_ui()
        self.apply_theme()
        self.load_tasks_to_ui()
        self.start_active_save_timer()
        self.show()
        self.destroyed.connect(self.on_window_destroyed)
        set_dark_title_bar_qt(self)
        self.update_datetime()

    def update_datetime(self):
        now = datetime.datetime.now()
        days = {
            0: "Понедельник",
            1: "Вторник",
            2: "Среда",
            3: "Четверг",
            4: "Пятница",
            5: "Суббота",
            6: "Воскресенье"
        }
        day_name = days[now.weekday()]
        self.datetime_label.setText(now.strftime("%d.%m.%Y %H:%M:%S") + " " + day_name)
        QTimer.singleShot(1000, self.update_datetime)

    def on_window_move_or_resize(self, event):
        app_config['USER']['window_x'] = str(self.x())
        app_config['USER']['window_y'] = str(self.y())
        app_config['USER']['window_width'] = str(self.width())
        app_config['USER']['window_height'] = str(self.height())
        save_settings(app_config)
        if isinstance(event, PyQt6.QtGui.QMoveEvent):
            super().moveEvent(event)
        else:
            super().resizeEvent(event)

    def start_active_save_timer(self):
        global active_save_timer
        if active_save_timer:
            return
        def save_loop():
            save_active_tasks()
            active_save_timer = QTimer.singleShot(10000, save_loop)
        save_loop()

    def init_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        top_layout = QHBoxLayout()
        self.datetime_label = QLabel()
        self.datetime_label.setStyleSheet("font-size: 12px; color: #888;")
        top_layout.addWidget(self.datetime_label)
        top_layout.addStretch()
        log_folder_btn = QPushButton("📂")
        log_folder_btn.setFixedHeight(26)
        log_folder_btn.clicked.connect(self.open_log_folder)
        log_folder_btn.setToolTip("Открыть папку с логами")
        top_layout.addWidget(log_folder_btn)
        change_folder_btn = QPushButton("⚙️")
        change_folder_btn.setFixedHeight(26)
        change_folder_btn.clicked.connect(self.change_save_folder)
        change_folder_btn.setToolTip("Изменить папку для отчетов")
        top_layout.addWidget(change_folder_btn)
        clear_logs_btn = QPushButton("🗑️")
        clear_logs_btn.setFixedHeight(26)
        clear_logs_btn.clicked.connect(self.clear_all_logs)
        clear_logs_btn.setToolTip("Очистить все логи и задачи")
        top_layout.addWidget(clear_logs_btn)
        main_layout.addLayout(top_layout)
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.tasks_container = QWidget()
        self.tasks_layout = QVBoxLayout(self.tasks_container)
        self.tasks_layout.addStretch()
        self.tasks_layout.setSpacing(5)
        scroll_area.setWidget(self.tasks_container)
        main_layout.addWidget(scroll_area)
        input_form = QFrame()
        input_form.setFrameShape(QFrame.Shape.StyledPanel)
        form_layout = QVBoxLayout(input_form)
        self.task_input = QLineEdit()
        self.task_input.setPlaceholderText("Задача")
        self.task_input.returnPressed.connect(self.handle_enter_pressed)
        self.link_input = QLineEdit()
        self.link_input.setPlaceholderText("Ссылка")
        self.link_input.returnPressed.connect(self.handle_enter_pressed)
        self.time_input = QLineEdit()
        self.time_input.setPlaceholderText("Время (минуты или 13:24-16:39)")
        self.time_input.returnPressed.connect(self.handle_enter_pressed)
        form_layout.addWidget(self.task_input)
        form_layout.addWidget(self.link_input)
        form_layout.addWidget(self.time_input)
        btn_layout = QHBoxLayout()
        add_btn = QPushButton("✅ Добавить задачу")
        add_btn.clicked.connect(self.add_task)
        stats_btn = QPushButton("📊 Статистика")
        stats_btn.clicked.connect(self.show_stats)
        save_btn = QPushButton("💾 Сохранить в Excel")
        save_btn.clicked.connect(self.save_excel_gui)
        quit_btn = QPushButton("🚪 Завершить день")
        quit_btn.clicked.connect(self.quit_app)
        btn_layout.addWidget(add_btn)
        btn_layout.addWidget(stats_btn)
        btn_layout.addWidget(save_btn)
        btn_layout.addWidget(quit_btn)
        form_layout.addLayout(btn_layout)
        main_layout.addWidget(input_form)

    def handle_enter_pressed(self):
        sender = self.sender()
        if sender == self.task_input:
            self.add_task()
        elif sender == self.link_input:
            self.add_task()
        elif sender == self.time_input:
            self.add_task()

    def apply_theme(self):
        old_move_event = self.moveEvent
        old_resize_event = self.resizeEvent
        self.moveEvent = lambda event: super(MainWindow, self).moveEvent(event)
        self.resizeEvent = lambda event: super(MainWindow, self).resizeEvent(event)
        try:
            palette = self.palette()
            palette.setColor(QPalette.ColorRole.Window, QColor(30, 30, 30))
            palette.setColor(QPalette.ColorRole.WindowText, QColor(220, 220, 220))
            palette.setColor(QPalette.ColorRole.Base, QColor(40, 40, 40))
            palette.setColor(QPalette.ColorRole.AlternateBase, QColor(50, 50, 50))
            palette.setColor(QPalette.ColorRole.ToolTipBase, QColor(25, 25, 25))
            palette.setColor(QPalette.ColorRole.ToolTipText, QColor(220, 220, 220))
            palette.setColor(QPalette.ColorRole.Text, QColor(220, 220, 220))
            palette.setColor(QPalette.ColorRole.Button, QColor(50, 50, 50))
            palette.setColor(QPalette.ColorRole.ButtonText, QColor(220, 220, 220))
            palette.setColor(QPalette.ColorRole.BrightText, QColor(255, 0, 0))
            palette.setColor(QPalette.ColorRole.Link, QColor(42, 130, 218))
            palette.setColor(QPalette.ColorRole.Highlight, QColor(42, 130, 218))
            palette.setColor(QPalette.ColorRole.HighlightedText, QColor(30, 30, 30))
            self.setStyleSheet(f"""
                QMainWindow, QWidget {{ background-color: #1e1e1e; color: #dcdcdc; }}
                QLineEdit, QTextEdit {{ background-color: #2a2a2a; color: #dcdcdc; border: 1px solid #555; padding: 5px; }}
                QPushButton {{ 
                    background-color: #3a3a3a; 
                    color: #dcdcdc; 
                    border: 1px solid #555; 
                    padding: 4px 8px;
                    margin: 2px;
                    border-radius: 4px;
                }}
                QPushButton:hover {{ background-color: #4a4a4a; }}
                QScrollArea {{ border: none; }}
                QLabel {{ color: #dcdcdc; }}
                QComboBox {{ background-color: #2a2a2a; color: #dcdcdc; border: 1px solid #555; }}
                /* === Стиль скроллбара === */
                QScrollBar:vertical {{
                    background: transparent;
                    width: 8px;
                    margin: 0px 0px 0px 0px;
                    border: none;
                }}
                QScrollBar::handle:vertical {{
                    background: rgba(100, 100, 100, 150);
                    border-radius: 4px;
                    min-height: 20px;
                }}
                QScrollBar::handle:vertical:hover {{
                    background: rgba(120, 120, 120, 200);
                }}
                /* Скрываем кнопки прокрутки */
                QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                    height: 0px;
                    width: 0px;
                    background: none;
                }}
                QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {{
                    background: none;
                }}
            """)
        finally:
            self.moveEvent = old_move_event
            self.resizeEvent = old_resize_event

    def reload_all_task_widgets(self):
        for i in reversed(range(self.tasks_layout.count() - 1)):
            widget = self.tasks_layout.itemAt(i).widget()
            if widget:
                widget.deleteLater()
        self.load_tasks_to_ui()

    def load_tasks_to_ui(self):
        for i in reversed(range(self.tasks_layout.count() - 1)):
            widget = self.tasks_layout.itemAt(i).widget()
            if widget:
                widget.deleteLater()
        self.tasks_widgets.clear()
        for idx, task in enumerate(tasks):
            task_widget = TaskItemWidget(task, idx, self)
            insert_pos = max(0, self.tasks_layout.count() - 1)
            self.tasks_layout.insertWidget(insert_pos, task_widget)
            self.tasks_widgets.append(task_widget)
        scroll_area = self.findChild(QScrollArea)
        if scroll_area:
            scroll_bar = scroll_area.verticalScrollBar()
            QTimer.singleShot(50, lambda: scroll_bar.setValue(scroll_bar.maximum()))

    def add_task(self):
        task_text = self.task_input.text().strip()
        if not task_text:
            QMessageBox.warning(self, "Ошибка", "Введите задачу!")
            return
        is_completed = task_text.startswith("!")
        if is_completed:
            task_text = task_text[1:].strip()
        task_text, extracted_link = extract_markdown_links(task_text)
        link = self.link_input.text().strip() or extracted_link
        time_str = self.time_input.text().strip()
        start_timestamp = time.time()
        if time_str:
            if "-" in time_str:
                minutes = parse_time_range(time_str)
                try:
                    start_str, end_str = time_str.split("-")
                    h1, m1 = map(int, start_str.split(":"))
                    h2, m2 = map(int, end_str.split(":"))
                    today = datetime.date.today()
                    start_dt = datetime.datetime(today.year, today.month, today.day, h1, m1)
                    end_dt = datetime.datetime(today.year, today.month, today.day, h2, m2)
                    if end_dt < start_dt:
                        end_dt += datetime.timedelta(days=1)
                    start_timestamp = start_dt.timestamp()
                    end_timestamp = end_dt.timestamp()
                except:
                    end_timestamp = start_timestamp + (minutes * 60)
            else:
                try:
                    minutes = float(time_str)
                except:
                    QMessageBox.warning(self, "Ошибка", "Неверный формат времени. Используйте число минут или формат 13:24-16:39")
                    return
                end_timestamp = start_timestamp + (minutes * 60)
            hours_hundredths = round(minutes / 60, 2)
            display_time_str = "<1 минуты" if minutes < 1 else f"{int(minutes)}:{int((minutes%1)*60):02d}"
            if "-" in time_str:
                display_time_str = time_str
            tasks.append({
                "id": str(uuid.uuid4()),
                "task": task_text,
                "link": link,
                "time_str": display_time_str,
                "minutes": round(minutes, 2),
                "hours_hundredths": hours_hundredths,
                "start_timestamp": start_timestamp,
                "end_timestamp": end_timestamp
            })
        else:
            tasks.append({
                "id": str(uuid.uuid4()),
                "task": task_text,
                "link": link,
                "timer_start": time.time(),
                "start_timestamp": start_timestamp,
                "minutes": 0.0,
                "hours_hundredths": 0.0,
                "time_str": "0:00"
            })
        save_backup()
        new_index = len(tasks) - 1
        task_widget = TaskItemWidget(tasks[new_index], new_index, self)
        insert_pos = max(0, self.tasks_layout.count() - 1)
        self.tasks_layout.insertWidget(insert_pos, task_widget)
        self.tasks_widgets.append(task_widget)
        scroll_area = self.findChild(QScrollArea)
        if scroll_area:
            scroll_bar = scroll_area.verticalScrollBar()
            QTimer.singleShot(50, lambda: scroll_bar.setValue(scroll_bar.maximum()))
        self.task_input.clear()
        self.link_input.clear()
        self.time_input.clear()

    def edit_task(self, index):
        if index < 0 or index >= len(tasks):
            return
        task = tasks[index]
        dialog = EditTaskDialog(task, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            new_data = dialog.get_data()
            if new_data:
                tasks[index].update(new_data)
                save_backup()
                self.load_tasks_to_ui()

    def delete_task(self, index):
        if index < 0 or index >= len(tasks):
            return
        reply = QMessageBox.question(self, "Подтверждение", "Вы уверены, что хотите удалить эту задачу?",
                                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            tasks.pop(index)
            save_backup()
            self.load_tasks_to_ui()

    def show_stats(self):
        total_tasks = len(tasks)
        total_minutes = round(sum(t.get('minutes', 0) for t in tasks), 2)
        total_hours = round(sum(t.get('hours_hundredths', 0) for t in tasks), 2)
        
        link_minutes = {}
        for t in tasks:
            if t.get('link'):
                d = parse_domain(t['link'])
                link_minutes[d] = link_minutes.get(d, 0) + t.get('minutes', 0)
        top_links = sorted(link_minutes.items(), key=lambda x: x[1], reverse=True)[:3]
        
        task_minutes = {}
        for t in tasks:
            task_minutes[t['task']] = task_minutes.get(t['task'], 0) + t.get('minutes', 0)
        top_tasks = sorted(task_minutes.items(), key=lambda x: x[1], reverse=True)[:3]
        
        stats_parts = []
        stats_parts.append("<b>📊 ОБЩАЯ СТАТИСТИКА</b>")
        stats_parts.append(f"  Всего задач: <b>{total_tasks}</b>")
        stats_parts.append(f"  Общее время: <b>{total_minutes} мин</b> / <b>{total_hours} ч</b>")
        
        if top_links:
            stats_parts.append("")
            stats_parts.append("<b>ТОП-3 ССЫЛОК ПО ВРЕМЕНИ</b>")
            for domain, mins in top_links:
                stats_parts.append(f"  🔗 {domain}: {round(mins, 2)} мин")
        
        if top_tasks:
            stats_parts.append("")
            stats_parts.append("<b>ТОП-3 ЗАДАЧ ПО ВРЕМЕНИ</b>")
            for name, mins in top_tasks:
                stats_parts.append(f"  📝 {name}: {round(mins, 2)} мин")
        
        stats_text = "<br>".join(stats_parts)
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("Статистика")
        msg_box.setTextFormat(Qt.TextFormat.RichText)
        msg_box.setText(stats_text)
        msg_box.exec()

    def save_excel_gui(self):
        global SAVE_DIR
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Сохранить отчет",
            os.path.join(SAVE_DIR, f"{datetime.date.today().strftime('%Y-%m-%d')}.xlsx"),
            "Excel Files (*.xlsx);;All Files (*)"
        )
        if file_path:
            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Tasks"
                ws.append(["Задача", "Ссылка (домен)", "Время (мин:сек)", "Время (часы в сотых)"])
                total_minutes = 0.0
                total_hours_hundredths = 0.0
                keys = [key_for_group(t) for t in tasks]
                dup_keys = {k for k, c in Counter(keys).items() if c > 1}
                row_to_key = {}
                for t in tasks:
                    domain = parse_domain(t['link']) if t['link'] else ""
                    row_index = ws.max_row + 1
                    ws.cell(row=row_index, column=1, value=t['task'])
                    if t['link']:
                        link_cell = ws.cell(row=row_index, column=2, value=domain)
                        link_cell.hyperlink = t['link']
                        link_cell.style = "Hyperlink"
                    else:
                        ws.cell(row=row_index, column=2, value="")
                    ws.cell(row=row_index, column=3, value=t['time_str'])
                    ws.cell(row=row_index, column=4, value=t['hours_hundredths'])
                    k = key_for_group(t)
                    if k in dup_keys:
                        row_to_key[row_index] = k
                    total_minutes += t['minutes']
                    total_hours_hundredths += t['hours_hundredths']
                n_tasks_rows = len(tasks) + 1
                ws.append([])
                ws.append(["ИТОГО", "", f"{round(total_minutes, 2)} мин", f"{round(total_hours_hundredths, 2)} ч"])
                grouped = group_tasks()
                if grouped:
                    ws.append([])
                    ws.append(["СВОДКА (объединено по совпадению ссылки или названия)"])
                    for gt in grouped:
                        ws.append([
                            gt["task"],
                            parse_domain(gt["link"]) if gt["link"] else "",
                            f"{round(gt['minutes'], 2)} мин",
                            f"{round(gt['hours_hundredths'], 2)} ч"
                        ])
                style_tasks_and_summary(
                    ws=ws,
                    n_tasks_rows=n_tasks_rows,
                    grouped=grouped,
                    group_key_to_rows={"row_to_key": row_to_key}
                )
                wb.save(file_path)
                new_save_dir = os.path.dirname(file_path)
                if new_save_dir != SAVE_DIR:
                    SAVE_DIR = new_save_dir
                    app_config['USER']['save_dir'] = SAVE_DIR
                    save_settings(app_config)
                QMessageBox.information(self, "Успех", f"Файл успешно сохранён:\n{file_path}")
                reply = QMessageBox.question(self, "Открыть папку?", "Открыть папку с файлом?",
                                            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
                if reply == QMessageBox.StandardButton.Yes:
                    if sys.platform == "win32":
                        os.startfile(os.path.dirname(file_path))
                    elif sys.platform == "darwin":
                        os.system(f'open "{os.path.dirname(file_path)}"')
                    else:
                        os.system(f'xdg-open "{os.path.dirname(file_path)}"')
            except PermissionError:
                QMessageBox.critical(self, "Ошибка", "Файл занят другим процессом. Пожалуйста, закройте файл и попробуйте снова.")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить файл:\n{str(e)}")

    def open_log_folder(self):
        if os.path.exists(LOG_DIR):
            if sys.platform == "win32":
                os.startfile(LOG_DIR)
            elif sys.platform == "darwin":
                os.system(f'open "{LOG_DIR}"')
            else:
                os.system(f'xdg-open "{LOG_DIR}"')
        else:
            QMessageBox.warning(self, "Папка не найдена", "Временная папка с логами ещё не создана.")

    def change_save_folder(self):
        global SAVE_DIR, app_config
        folder = QFileDialog.getExistingDirectory(
            self,
            "Выберите папку для сохранения отчётов Excel",
            SAVE_DIR
        )
        if folder:
            SAVE_DIR = folder
            app_config['USER']['save_dir'] = SAVE_DIR
            save_settings(app_config)
            QMessageBox.information(self, "Успех", f"Папка сохранения изменена на:\n{SAVE_DIR}")

    def clear_all_logs(self):
        reply = QMessageBox.question(
            self,
            "Подтверждение",
            "Вы уверены, что хотите удалить ВСЕ логи и сбросить список задач?\n"
            "Это действие необратимо.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        try:
            if os.path.exists(LOG_DIR):
                for file in os.listdir(LOG_DIR):
                    if file.endswith(".txt"):
                        os.remove(os.path.join(LOG_DIR, file))
            if os.path.exists(SAVE_DIR):
                for file in os.listdir(SAVE_DIR):
                    if file.endswith(".xlsx") and file.startswith(datetime.date.today().strftime("%Y-%m-%d")):
                        os.remove(os.path.join(SAVE_DIR, file))
            tasks.clear()
            save_backup()
            self.load_tasks_to_ui()
            QMessageBox.information(self, "Очистка", "Все логи и задачи удалены.")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось очистить логи:\n{str(e)}")

    def quit_app(self):
        reply = QMessageBox.question(
            self,
            "Завершение дня",
            "Хотите сохранить отчет перед выходом?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No | QMessageBox.StandardButton.Cancel
        )
        if reply == QMessageBox.StandardButton.Yes:
            self.save_excel_gui()
            save_backup()
            if os.path.exists(active_tasks_file):
                try:
                    os.remove(active_tasks_file)
                except Exception as e:
                    print(Fore.YELLOW + f"Не удалось удалить временный файл: {e}")
            QApplication.quit()
        elif reply == QMessageBox.StandardButton.No:
            save_backup()
            if os.path.exists(active_tasks_file):
                try:
                    pass
                    #os.remove(active_tasks_file)
                except Exception as e:
                    print(Fore.YELLOW + f"Не удалось удалить временный файл: {e}")
            QApplication.quit()
        else:
            pass

    def on_window_destroyed(self):
        save_settings(app_config)

    def moveEvent(self, event):
        self.save_window_state()
        super().moveEvent(event)

    def resizeEvent(self, event):
        self.save_window_state()
        super().resizeEvent(event)

    def save_window_state(self):
        try:
            app_config['USER']['window_x'] = str(self.x())
            app_config['USER']['window_y'] = str(self.y())
            app_config['USER']['window_width'] = str(self.width())
            app_config['USER']['window_height'] = str(self.height())
            save_settings(app_config)
        except Exception as e:
            import traceback
            with open("error_log.txt", "a", encoding='utf-8') as logf:
                logf.write(f"[ERROR_save_window_state] {e}\n{traceback.format_exc()}\n")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())