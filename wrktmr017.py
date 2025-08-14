# v0.1.7 — Журнал задач с доменами, кликабельными ссылками, объединением и редактированием

import os
import sys
import time
import datetime
import signal
import re
from urllib.parse import urlparse
import openpyxl
from colorama import init, Fore, Style

init(autoreset=True)

# ======================
# Настройки
# ======================

def get_log_dir():
    desktop_dir = os.path.join(os.path.expanduser("~"), "Desktop")
    today = datetime.date.today()
    return os.path.join(desktop_dir, "logs", str(today.year), f"{today.month:02d}", f"{today.day:02d}")

LOG_DIR = get_log_dir()
os.makedirs(LOG_DIR, exist_ok=True)

tasks = []
start_time = None
last_task_time = 0

# ======================
# Служебные функции
# ======================

def signal_handler(sig, frame):
    print(Fore.YELLOW + "\nОбнаружено Ctrl+C! Сохраняем данные...")
    save_backup()
    save_excel()
    print(Fore.MAGENTA + "Данные сохранены. Программа завершена.")
    sys.exit(0)

signal.signal(signal.SIGINT, signal_handler)

def parse_domain(url):
    """Возвращает только домен из ссылки"""
    try:
        parsed = urlparse(url)
        return parsed.netloc or url
    except:
        return url

def link_clickable(url, text=None):
    """Делает ссылку кликабельной в консоли"""
    if not url:
        return ""
    if text is None:
        text = url
    return f"\033]8;;{url}\033\\{text}\033]8;;\033\\"

def load_backup():
    today = datetime.date.today().strftime("%Y-%m-%d")
    log_file = os.path.join(LOG_DIR, f"{today}.txt")
    if os.path.exists(log_file):
        with open(log_file, "r", encoding="utf-8") as f:
            for line in f:
                try:
                    task, link, time_str, hours_hundredths = line.strip().split(" | ")
                    if time_str == "<1 минуты":
                        minutes = 0.5
                    else:
                        mins, secs = map(int, time_str.split(":"))
                        minutes = mins + secs / 60
                    tasks.append({
                        "task": task,
                        "link": link,
                        "time_str": time_str,
                        "minutes": round(minutes, 2),
                        "hours_hundredths": float(hours_hundredths)
                    })
                except Exception as e:
                    print(Fore.RED + f"Ошибка чтения backup: {e}")
        if tasks:
            print(Fore.GREEN + f"Загружено {len(tasks)} задач из backup: {log_file}")

def clear_console():
    os.system("cls" if os.name == "nt" else "clear")

def save_backup():
    today = datetime.date.today().strftime("%Y-%m-%d")
    log_file = os.path.join(LOG_DIR, f"{today}.txt")
    with open(log_file, "w", encoding="utf-8") as f:
        for t in tasks:
            f.write(f"{t['task']} | {t['link']} | {t['time_str']} | {t['hours_hundredths']}\n")

def group_tasks():
    """Объединяет задачи с одинаковым названием или ссылкой"""
    grouped = {}
    for t in tasks:
        name_key = t["task"].strip().lower()
        link_key = t["link"].strip()
        key = name_key if name_key else link_key
        # Если нет названия, используем ссылку как ключ, и наоборот
        if key not in grouped:
            grouped[key] = {
                "task": t["task"],
                "link": t["link"],
                "minutes": 0.0,
                "hours_hundredths": 0.0,
                "count": 0
            }
        grouped[key]["minutes"] += t["minutes"]
        grouped[key]["hours_hundredths"] += t["hours_hundredths"]
        grouped[key]["count"] += 1
    # Оставляем только те, где было 2+ одинаковых
    return [g for g in grouped.values() if g["count"] > 1]

def save_excel():
    today = datetime.date.today().strftime("%Y-%m-%d")
    file_name = os.path.join(LOG_DIR, f"{today}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tasks"

    ws.append(["Задача", "Ссылка", "Время (мин:сек)", "Время (часы в сотых)"])

    total_minutes = 0.0
    total_hours_hundredths = 0.0

    for t in tasks:
        domain = parse_domain(t['link']) if t['link'] else ""
        row = [t['task']]
        if t['link']:
            link_cell = ws.cell(row=ws.max_row + 1, column=2, value=domain)
            link_cell.hyperlink = t['link']
        else:
            ws.cell(row=ws.max_row + 1, column=2, value="")
        ws.cell(row=ws.max_row, column=1, value=t['task'])
        ws.cell(row=ws.max_row, column=3, value=t['time_str'])
        ws.cell(row=ws.max_row, column=4, value=t['hours_hundredths'])

        total_minutes += t['minutes']
        total_hours_hundredths += t['hours_hundredths']

    ws.append([])
    ws.append(["ИТОГО", "", f"{round(total_minutes, 2)} мин", f"{round(total_hours_hundredths, 2)} ч"])

    grouped = group_tasks()
    if grouped:
        ws.append([])
        ws.append(["СВОДКА (объединено по совпадению названия или ссылки)"])
        for gt in grouped:
            ws.append([
                gt["task"],
                parse_domain(gt["link"]),
                f"{round(gt['minutes'], 2)} мин",
                f"{round(gt['hours_hundredths'], 2)} ч"
            ])

    try:
        wb.save(file_name)
        print(Fore.GREEN + f"Excel-отчёт сохранён: {file_name}")
    except Exception as e:
        print(Fore.RED + f"Ошибка при сохранении Excel: {e}")

def display_tasks():
    clear_console()
    print(Fore.CYAN + "=== Журнал задач ===")
    for idx, t in enumerate(tasks, start=1):
        domain = parse_domain(t['link']) if t['link'] else ""
        link_part = ""
        if t['link']:
            link_part = f" [{link_clickable(t['link'], domain)}]"
        print(
            Fore.YELLOW + f"{idx}. {t['task']}{link_part}" +
            Fore.WHITE + f" — {t['time_str']} ({t['hours_hundredths']} ч)"
        )

def edit_task(index):
    try:
        t = tasks[index]
        print(Fore.CYAN + f"Редактирование задачи #{index+1}")
        new_task = input(f"Текст задачи ({t['task']}): ").strip() or t['task']
        new_link = input(f"Ссылка ({t['link']}): ").strip() or t['link']
        new_time = input(f"Время (минуты, текущее {t['minutes']}): ").strip()

        if new_time:
            try:
                minutes = float(new_time)
            except:
                print(Fore.RED + "Неверный ввод времени, оставлено старое значение")
                minutes = t['minutes']
        else:
            minutes = t['minutes']

        tasks[index] = {
            "task": new_task,
            "link": new_link,
            "minutes": minutes,
            "hours_hundredths": round(minutes / 60, 2),
            "time_str": "<1 минуты" if minutes < 1 else f"{int(minutes)}:{int((minutes%1)*60):02d}"
        }
        save_backup()
        display_tasks()
    except IndexError:
        print(Fore.RED + "Нет задачи с таким номером.")

# ======================
# Запуск
# ======================

load_backup()
display_tasks()

print(Fore.GREEN + "Вводите задачи (! — уже выполненная). Команды: :h — завершить день, :s — сохранить, :q — выйти, :e1 — редактировать задачу №1.")

while True:
    try:
        task_input = input(Fore.CYAN + "\nВведите задачу: ").strip()

        if re.match(r"^:e\d+$", task_input) or re.match(r"^:edit\d+$", task_input):
            idx = int(re.findall(r"\d+", task_input)[0]) - 1
            edit_task(idx)
            continue

        if task_input.lower() in [":home", ":h", ":q", ":quit", ":exit"]:
            save_excel()
            print(Fore.MAGENTA + "День завершён.")
            break
        elif task_input.lower() in [":s", ":save"]:
            save_excel()
            continue
        elif not task_input:
            print(Fore.RED + "Пожалуйста, введите задачу или команду.")
            continue

        is_completed = task_input.startswith("!")
        task = task_input[1:] if is_completed else task_input
        link = input(Fore.CYAN + "Введите ссылку (если есть): ").strip()

        if is_completed:
            while True:
                try:
                    minutes = float(input(Fore.CYAN + "Время (в минутах): ").strip())
                    break
                except ValueError:
                    print(Fore.RED + "Введите число!")
        else:
            start_time = time.time()
            input(Fore.CYAN + "Нажмите Enter, когда задача завершена...")
            elapsed = time.time() - start_time
            minutes = elapsed / 60

        hours_hundredths = round(minutes / 60, 2)
        time_str = "<1 минуты" if minutes < 1 else f"{int(minutes)}:{int((minutes%1)*60):02d}"

        tasks.append({
            "task": task,
            "link": link,
            "time_str": time_str,
            "minutes": round(minutes, 2),
            "hours_hundredths": hours_hundredths
        })

        save_backup()
        display_tasks()

    except KeyboardInterrupt:
        signal_handler(None, None)
